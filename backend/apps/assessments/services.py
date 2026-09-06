import io
import random
import secrets
import uuid
from datetime import timedelta
from typing import List, Dict, Any, Optional, Tuple
from django.db import models, transaction
from django.db.models import Q
from django.utils import timezone
from django.core.exceptions import PermissionDenied, ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

from apps.accounts.models import AuditLog, User, Role, Section, StudentProfile
from apps.accounts.services import AuditService
from apps.questions.models import QuestionVersion, VersionStatus, QuestionType
from .models import (
    Assessment,
    AssessmentStatus,
    AssessmentAssignment,
    AssignmentStatus,
    AssessmentQuestion,
    AssessmentSnapshot,
    AssessmentSnapshotQuestion,
    TestAttempt,
    AttemptStatus,
    AttemptAnswer,
)


class RandomizationService:
    """
    Deterministic seeded randomization engine for question and option presentation.
    """
    @staticmethod
    def generate_seed() -> str:
        return secrets.token_hex(16)

    @classmethod
    def randomize_question_order(cls, seed: str, question_ids: List[str], randomize: bool = True) -> List[str]:
        q_list = list(question_ids)
        if not randomize or len(q_list) <= 1:
            return q_list
        rng = random.Random(f"{seed}_questions")
        rng.shuffle(q_list)
        return q_list

    @classmethod
    def randomize_options(cls, seed: str, question_id: str, options: List[Dict[str, Any]], randomize: bool = True) -> List[str]:
        """
        Returns list of option IDs in randomized order.
        """
        opt_ids = [opt['id'] for opt in options if 'id' in opt]
        if not randomize or len(opt_ids) <= 1:
            return opt_ids
        rng = random.Random(f"{seed}_opts_{question_id}")
        rng.shuffle(opt_ids)
        return opt_ids


class AttemptTimerService:
    """
    Server-authoritative timer and deadline calculations.
    """
    @staticmethod
    def compute_expiry(started_at, duration_minutes: int, assessment_end) -> Any:
        duration_expiry = started_at + timedelta(minutes=duration_minutes)
        # Effective deadline is MIN(start + duration, assessment.end_datetime)
        if assessment_end and assessment_end < duration_expiry:
            return assessment_end
        return duration_expiry

    @staticmethod
    def get_remaining_seconds(attempt: TestAttempt) -> int:
        if attempt.status in [AttemptStatus.SUBMITTED, AttemptStatus.EXPIRED, AttemptStatus.CANCELLED]:
            return 0
        if not attempt.expires_at:
            return 0
        now = timezone.now()
        remaining = int((attempt.expires_at - now).total_seconds())
        return max(0, remaining)

    @staticmethod
    def is_expired(attempt: TestAttempt, grace_seconds: int = 5) -> bool:
        if not attempt.expires_at:
            return False
        now = timezone.now()
        # Allows 5-second grace strictly for network flight, never extends exam writing
        return now > (attempt.expires_at + timedelta(seconds=grace_seconds))

    @classmethod
    def check_and_expire_attempt_if_needed(cls, attempt: TestAttempt) -> bool:
        if attempt.status == AttemptStatus.IN_PROGRESS and cls.is_expired(attempt, grace_seconds=0):
            attempt.status = AttemptStatus.EXPIRED
            attempt.save(update_fields=['status', 'updated_at'])
            AuditService.log(
                action="ATTEMPT_EXPIRED",
                actor=attempt.student,
                target_type="TestAttempt",
                target_id=str(attempt.id),
                metadata={
                    "assessment_id": str(attempt.assessment_id),
                    "attempt_number": attempt.attempt_number,
                    "reason": "Timer exceeded server deadline."
                }
            )
            return True
        return False

    @classmethod
    def authorize_pause(
        cls,
        attempt: TestAttempt,
        actor: Optional[User] = None
    ) -> bool:
        """
        Server-authoritative validation determining if an attempt is eligible for pause.
        Strictly requires IN_PROGRESS status and non-expired server timeline.
        """
        if attempt.status != AttemptStatus.IN_PROGRESS:
            raise DRFValidationError({
                "status": f"Cannot pause attempt in status {attempt.status}. Only IN_PROGRESS attempts may be paused."
            })

        # Check if already expired according to server authority
        if cls.check_and_expire_attempt_if_needed(attempt):
            raise DRFValidationError({"status": "Cannot pause an expired attempt."})

        # Check assessment hard window
        now = timezone.now()
        if attempt.assessment.end_datetime and now >= attempt.assessment.end_datetime:
            raise DRFValidationError({"schedule": "Cannot pause attempt: assessment end datetime has passed."})

        return True

    @classmethod
    @transaction.atomic
    def apply_authorized_pause(
        cls,
        attempt: TestAttempt,
        pause_duration_seconds: int,
        actor: Optional[User] = None,
        request=None
    ) -> TestAttempt:
        """
        Server-authoritative operation extending attempt.expires_at by the authorized
        pause duration, strictly clamped to assessment.end_datetime as an absolute ceiling.
        """
        if attempt.status != AttemptStatus.IN_PROGRESS:
            raise DRFValidationError({
                "status": f"Cannot apply pause to attempt in status {attempt.status}."
            })

        if pause_duration_seconds <= 0:
            raise DRFValidationError({
                "duration": "Pause duration must be a positive integer in seconds."
            })

        if attempt.expires_at:
            new_expiry = attempt.expires_at + timedelta(seconds=pause_duration_seconds)
            # Hard invariant: effective expiry can never exceed assessment.end_datetime
            if attempt.assessment.end_datetime and new_expiry > attempt.assessment.end_datetime:
                new_expiry = attempt.assessment.end_datetime

            attempt.expires_at = new_expiry
            attempt.save(update_fields=['expires_at', 'updated_at'])

            AuditService.log(
                action="ATTEMPT_TIMER_PAUSE_APPLIED",
                actor=actor or attempt.student,
                target_type="TestAttempt",
                target_id=str(attempt.id),
                metadata={
                    "assessment_id": str(attempt.assessment_id),
                    "pause_duration_seconds": pause_duration_seconds,
                    "new_expires_at": attempt.expires_at.isoformat(),
                    "assessment_end": attempt.assessment.end_datetime.isoformat() if attempt.assessment.end_datetime else None
                },
                request=request
            )

        return attempt



class AssessmentSnapshotService:
    """
    Constructs and permanently freezes self-contained AssessmentSnapshot and AssessmentSnapshotQuestion records.
    """
    @classmethod
    @transaction.atomic
    def create_snapshot(cls, assessment: Assessment, actor: Optional[User] = None, request=None) -> AssessmentSnapshot:
        # Prevent re-snapshotting
        if hasattr(assessment, 'snapshot') and assessment.snapshot is not None:
            return assessment.snapshot

        assessment_questions = list(assessment.assessment_questions.select_related('question_version').order_by('order'))
        if not assessment_questions:
            raise DRFValidationError({"questions": "Assessment must contain at least one question before publishing."})

        # Calculate student-safe snapshot bundle and server-only evaluation bundle
        student_questions_list = []
        server_eval_questions_map = {}

        # 1. Create base AssessmentSnapshot
        snapshot = AssessmentSnapshot.objects.create(
            assessment=assessment,
            version_number=1,
            snapshot_data={},
            server_evaluation_bundle={}
        )

        for aq in assessment_questions:
            qv = aq.question_version
            snapshot_q_id = str(qv.id)

            # Public test cases vs Hidden test cases
            public_coding_config = {}
            server_coding_eval = {}

            if qv.question_type == QuestionType.CODING and hasattr(qv, 'coding_config') and qv.coding_config:
                c_conf = qv.coding_config
                all_tcs = list(c_conf.test_cases.all())
                public_tcs = [
                    {
                        "input_data": tc.input_data,
                        "expected_output": tc.expected_output,
                        "points": tc.points,
                        "is_hidden": False,
                        "execution_order": tc.execution_order
                    }
                    for tc in all_tcs if not tc.is_hidden
                ]
                hidden_tcs = [
                    {
                        "input_data": tc.input_data,
                        "expected_output": tc.expected_output,
                        "points": tc.points,
                        "is_hidden": True,
                        "execution_order": tc.execution_order
                    }
                    for tc in all_tcs if tc.is_hidden
                ]

                public_coding_config = {
                    "problem_statement": c_conf.problem_statement,
                    "input_description": c_conf.input_description,
                    "output_description": c_conf.output_description,
                    "constraints": c_conf.constraints,
                    "allowed_languages": c_conf.allowed_languages,
                    "time_limit_ms": c_conf.time_limit_ms,
                    "memory_limit_mb": c_conf.memory_limit_mb,
                    "public_test_cases": public_tcs
                }
                server_coding_eval = {
                    "all_test_cases": public_tcs + hidden_tcs,
                    "allowed_languages": c_conf.allowed_languages,
                    "time_limit_ms": c_conf.time_limit_ms,
                    "memory_limit_mb": c_conf.memory_limit_mb,
                }

            sql_config_data = {}
            server_sql_eval = {}
            if qv.question_type == QuestionType.SQL and hasattr(qv, 'sql_config') and qv.sql_config:
                s_conf = qv.sql_config
                sql_config_data = {
                    "problem_statement": s_conf.problem_statement,
                    "schema_setup_sql": s_conf.schema_setup_sql,
                    "allowed_dialect": s_conf.allowed_dialect,
                    "time_limit_ms": s_conf.time_limit_ms
                }
                server_sql_eval = {
                    "expected_result_definition": s_conf.expected_result_definition,
                    "schema_setup_sql": s_conf.schema_setup_sql,
                    "allowed_dialect": s_conf.allowed_dialect,
                    "time_limit_ms": s_conf.time_limit_ms
                }

            # Negative points resolution
            neg_enabled = assessment.negative_marking_enabled and aq.negative_marking_enabled
            neg_pts = aq.negative_points if neg_enabled else 0

            # Tags
            tags_list = list(qv.tags.values_list('name', flat=True))

            # 2. Create AssessmentSnapshotQuestion entity
            snap_q = AssessmentSnapshotQuestion.objects.create(
                snapshot=snapshot,
                question_version=qv,
                snapshot_question_id=snapshot_q_id,
                order=aq.order,
                question_type=qv.question_type,
                title=qv.title,
                description=qv.description,
                instructions=qv.instructions,
                points=aq.points,
                negative_marking_enabled=neg_enabled,
                negative_points=neg_pts,
                difficulty=qv.difficulty,
                type_config=qv.type_config or {},
                coding_config=public_coding_config,
                sql_config=sql_config_data,
                tags=tags_list
            )

            student_questions_list.append({
                "snapshot_question_id": snapshot_q_id,
                "order": aq.order,
                "question_type": qv.question_type,
                "title": qv.title,
                "description": qv.description,
                "instructions": qv.instructions,
                "points": aq.points,
                "negative_marking_enabled": neg_enabled,
                "negative_points": neg_pts,
                "difficulty": qv.difficulty,
                "type_config": qv.type_config or {},
                "coding_config": public_coding_config,
                "sql_config": sql_config_data,
                "tags": tags_list
            })

            server_eval_questions_map[snapshot_q_id] = {
                "question_version_id": str(qv.id),
                "question_type": qv.question_type,
                "points": aq.points,
                "negative_marking_enabled": neg_enabled,
                "negative_points": neg_pts,
                "correct_type_config": qv.type_config or {},
                "server_coding_eval": server_coding_eval,
                "server_sql_eval": server_sql_eval,
            }

        # 3. Assemble top-level frozen bundles
        snapshot_data = {
            "assessment_id": str(assessment.id),
            "title": assessment.title,
            "description": assessment.description,
            "instructions": assessment.instructions,
            "duration_minutes": assessment.duration_minutes,
            "total_points": assessment.total_points,
            "passing_percentage": float(assessment.passing_percentage),
            "negative_marking_enabled": assessment.negative_marking_enabled,
            "attempt_limit": assessment.attempt_limit,
            "randomize_questions": assessment.randomize_questions,
            "randomize_options": assessment.randomize_options,
            "result_visibility": assessment.result_visibility,
            "start_datetime": assessment.start_datetime.isoformat(),
            "end_datetime": assessment.end_datetime.isoformat(),
            "questions": student_questions_list,
        }

        server_eval_bundle = {
            "assessment_id": str(assessment.id),
            "total_points": assessment.total_points,
            "questions_eval": server_eval_questions_map,
        }

        # Direct SQL update or bypassing save immutability for initial payload insert
        AssessmentSnapshot.objects.filter(pk=snapshot.pk).update(
            snapshot_data=snapshot_data,
            server_evaluation_bundle=server_eval_bundle
        )
        snapshot.refresh_from_db()

        AuditService.log(
            action="SNAPSHOT_CREATED",
            actor=actor or assessment.created_by,
            target_type="AssessmentSnapshot",
            target_id=str(snapshot.id),
            metadata={
                "assessment_id": str(assessment.id),
                "question_count": len(student_questions_list),
                "total_points": assessment.total_points
            },
            request=request
        )

        return snapshot


def resolve_student_users(student_ids: Any, active_only: bool = True) -> List[User]:
    """
    Centralized student identity resolver for the assessment domain.
    Resolves both canonical User.id and legacy/API StudentProfile.id into canonical User objects.
    Enforces Role.STUDENT, presence of StudentProfile, and active status.
    Raises DRFValidationError (HTTP 400) if any supplied ID cannot be resolved or is invalid.
    """
    if not student_ids:
        return []

    clean_ids = [str(sid).strip() for sid in student_ids if sid and str(sid).strip()]
    if not clean_ids:
        return []

    valid_uuid_ids = []
    malformed_ids = []
    for sid in clean_ids:
        try:
            uuid_obj = uuid.UUID(sid)
            valid_uuid_ids.append(str(uuid_obj))
        except (ValueError, AttributeError):
            malformed_ids.append(sid)

    if malformed_ids:
        raise DRFValidationError({
            "error": "INVALID_STUDENT_IDS",
            "invalid_ids": malformed_ids,
            "students": f"Account(s) {', '.join(malformed_ids)} cannot be targeted as a student (must be valid student IDs).",
            "detail": f"The following student ID(s) cannot be targeted as a student: {', '.join(malformed_ids)}"
        })

    candidate_users = list(
        User.objects.filter(
            Q(id__in=valid_uuid_ids) | Q(student_profile__id__in=valid_uuid_ids)
        )
        .select_related('student_profile', 'student_profile__section')
        .distinct()
    )

    resolved_id_map: Dict[str, User] = {}
    invalid_ids = []

    for sid in clean_ids:
        matched_user = None
        for u in candidate_users:
            profile_id = str(u.student_profile.id) if hasattr(u, 'student_profile') and u.student_profile else None
            if str(u.id) == sid or profile_id == sid:
                matched_user = u
                break

        if not matched_user:
            invalid_ids.append(sid)
        elif matched_user.role != Role.STUDENT:
            invalid_ids.append(sid)
        elif active_only and not matched_user.is_active:
            invalid_ids.append(sid)
        else:
            resolved_id_map[sid] = matched_user

    if invalid_ids:
        raise DRFValidationError({
            "error": "INVALID_STUDENT_IDS",
            "invalid_ids": invalid_ids,
            "students": f"Account(s) {', '.join(invalid_ids)} cannot be targeted as a student (must be active student accounts).",
            "detail": f"The following student ID(s) cannot be targeted as a student: {', '.join(invalid_ids)}"
        })

    # Return deduplicated canonical User objects preserving caller order
    seen_user_ids = set()
    canonical_users = []
    for sid in clean_ids:
        u = resolved_id_map[sid]
        if str(u.id) not in seen_user_ids:
            seen_user_ids.add(str(u.id))
            canonical_users.append(u)

    return canonical_users


class AssessmentAudienceService:
    """
    Authoritative domain service for resolving and configuring assessment audience targeting.
    Enforces that Section is purely a selection mechanism while AssessmentAssignment is authoritative access control.
    """
    @classmethod
    def resolve_audience(
        cls,
        assessment: Assessment,
        section_ids: Optional[List[str]] = None,
        student_ids: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Pure authoritative audience calculation.
        Resolves Section students UNION Individual students, deduplicated by User ID.
        Never creates persistent assignment records.
        """
        if section_ids is None:
            section_ids = [str(sid) for sid in assessment.target_sections.values_list('id', flat=True)]
        else:
            section_ids = [str(sid) for sid in section_ids if sid]

        if student_ids is None:
            target_ids = set(assessment.target_students.values_list('id', flat=True))
            assigned_ids = set(assessment.assignments.filter(status=AssignmentStatus.ASSIGNED).values_list('student_id', flat=True))
            student_ids = [str(uid) for uid in target_ids.union(assigned_ids)]
        else:
            student_ids = [str(uid) for uid in student_ids if uid]

        # 1. Resolve Section Students
        sections = list(Section.objects.filter(id__in=section_ids))
        section_students = list(User.objects.filter(
            role=Role.STUDENT,
            is_active=True,
            student_profile__isnull=False,
            student_profile__section__in=sections
        ).select_related('student_profile', 'student_profile__section').distinct())

        section_student_ids = {str(st.id) for st in section_students}

        sections_summary = []
        for sec in sections:
            count = sum(1 for st in section_students if st.student_profile and st.student_profile.section_id == sec.id)
            sections_summary.append({
                "id": str(sec.id),
                "code": sec.code,
                "name": sec.name,
                "is_active": sec.is_active,
                "student_count": count
            })

        # 2. Resolve Individually Targeted Students using canonical resolver
        individual_students = resolve_student_users(student_ids, active_only=True)
        individual_student_ids = {str(st.id) for st in individual_students}

        additional_students_summary = []
        for st in individual_students:
            profile = getattr(st, 'student_profile', None)
            additional_students_summary.append({
                "id": str(st.id),
                "user_id": str(st.id),
                "student_profile_id": str(profile.id) if profile else None,
                "email": st.email,
                "display_name": st.display_name,
                "roll_number": profile.roll_number if profile else "",
                "euid": profile.euid if profile else "",
                "section": profile.section.code if (profile and profile.section) else None
            })

        # 3. UNION and Deduplicate by Student User ID
        final_eligible_ids = sorted(list(section_student_ids.union(individual_student_ids)))
        overlap_ids = sorted(list(section_student_ids.intersection(individual_student_ids)))

        total_eligible = len(final_eligible_ids)
        return {
            "section_student_count": len(section_student_ids),
            "individual_student_count": len(individual_student_ids),
            "overlap_count": len(overlap_ids),
            "total_eligible": total_eligible,
            "total_eligible_count": total_eligible,
            "eligible_student_ids": final_eligible_ids,
            "sections": sections_summary,
            "additional_students": additional_students_summary,
            "students": additional_students_summary
        }

    @classmethod
    def configure_audience(
        cls,
        assessment: Assessment,
        section_ids: Optional[List[Any]] = None,
        student_ids: Optional[List[Any]] = None,
        actor: Optional[User] = None,
        request=None
    ) -> Dict[str, Any]:
        """
        Configures draft assessment audience.
        Rejects audience mutation on PUBLISHED or ARCHIVED assessments.
        """
        if assessment.status in [AssessmentStatus.PUBLISHED, AssessmentStatus.ARCHIVED]:
            raise PermissionDenied(
                "Published or archived assessments cannot change their target audience. "
                "Use assignment management to add or revoke individual access."
            )

        clean_section_ids = [str(sid) for sid in section_ids if sid] if section_ids is not None else []
        clean_student_ids = [str(uid) for uid in student_ids if uid] if student_ids is not None else []

        # Validate sections exist and are active
        sections = list(Section.objects.filter(id__in=clean_section_ids))
        if len(sections) != len(set(clean_section_ids)):
            raise DRFValidationError({"sections": "One or more selected sections do not exist."})
        for sec in sections:
            if not sec.is_active:
                raise DRFValidationError({"sections": f"Inactive section '{sec.code}' cannot be targeted."})

        # Validate and resolve individual students using canonical resolver
        students = resolve_student_users(clean_student_ids, active_only=True)

        with transaction.atomic():
            assessment.target_sections.set(sections)
            assessment.target_students.set(students)

            resolved = cls.resolve_audience(assessment)

            AuditService.log(
                action="ASSESSMENT_AUDIENCE_CONFIGURED",
                actor=actor,
                target_type="Assessment",
                target_id=str(assessment.id),
                metadata={
                    "title": assessment.title,
                    "section_codes": [s.code for s in sections],
                    "individual_student_count": len(students),
                    "total_eligible": resolved["total_eligible"]
                },
                request=request
            )

        return resolved


class AssessmentService:
    """
    Authoritative domain service for Assessment lifecycle, configuration, publishing, and assignments.
    """
    @classmethod
    @transaction.atomic
    def create_assessment(
        cls,
        title: str,
        *args,
        description: str = "",
        start_datetime = None,
        end_datetime = None,
        duration_minutes: int = 60,
        total_points: int = 0,
        created_by: Optional[User] = None,
        instructions: str = '',
        negative_marking_enabled: bool = False,
        attempt_limit: int = 1,
        randomize_questions: bool = False,
        randomize_options: bool = False,
        passing_percentage = None,
        result_visibility: str = 'AFTER_DEADLINE',
        request=None,
        **kwargs
    ) -> Assessment:
        if len(args) >= 1:
            description = args[0]
        if len(args) >= 2:
            start_datetime = args[1]
        if len(args) >= 3:
            end_datetime = args[2]
        if len(args) >= 4:
            duration_minutes = args[3]
        if len(args) >= 5:
            total_points = args[4]
        if len(args) >= 6:
            created_by = args[5]

        if start_datetime is None or end_datetime is None:
            raise DRFValidationError({"dates": "start_datetime and end_datetime are required."})
        if end_datetime <= start_datetime:
            raise DRFValidationError({"end_datetime": "End datetime must be strictly after start datetime."})
        if duration_minutes < 1:
            raise DRFValidationError({"duration_minutes": "Duration must be at least 1 minute."})
        if attempt_limit < 1:
            raise DRFValidationError({"attempt_limit": "Attempt limit must be at least 1."})

        assessment = Assessment.objects.create(
            title=title.strip(),
            description=(description or '').strip(),
            instructions=(instructions or '').strip(),
            start_datetime=start_datetime,
            end_datetime=end_datetime,
            duration_minutes=duration_minutes,
            total_points=total_points,
            passing_percentage=passing_percentage if passing_percentage is not None else 0.00,
            negative_marking_enabled=negative_marking_enabled,
            attempt_limit=attempt_limit,
            randomize_questions=randomize_questions,
            randomize_options=randomize_options,
            result_visibility=result_visibility,
            created_by=created_by,
            status=AssessmentStatus.DRAFT
        )

        AuditService.log(
            action="ASSESSMENT_CREATED",
            actor=created_by,
            target_type="Assessment",
            target_id=str(assessment.id),
            metadata={"title": assessment.title},
            request=request
        )
        return assessment

    @classmethod
    @transaction.atomic
    def update_draft_assessment(
        cls,
        assessment: Assessment,
        actor: User,
        title: Optional[str] = None,
        description: Optional[str] = None,
        instructions: Optional[str] = None,
        start_datetime = None,
        end_datetime = None,
        duration_minutes: Optional[int] = None,
        total_points: Optional[int] = None,
        negative_marking_enabled: Optional[bool] = None,
        attempt_limit: Optional[int] = None,
        randomize_questions: Optional[bool] = None,
        randomize_options: Optional[bool] = None,
        passing_percentage = None,
        result_visibility: Optional[str] = None,
        target_section_ids: Optional[List[Any]] = None,
        target_student_ids: Optional[List[Any]] = None,
        request=None
    ) -> Assessment:
        if assessment.status in [AssessmentStatus.PUBLISHED, AssessmentStatus.ARCHIVED]:
            raise PermissionDenied("Cannot edit a published or archived assessment.")

        if title is not None:
            assessment.title = title.strip()
        if description is not None:
            assessment.description = description.strip()
        if instructions is not None:
            assessment.instructions = instructions.strip()
        if start_datetime is not None:
            assessment.start_datetime = start_datetime
        if end_datetime is not None:
            assessment.end_datetime = end_datetime
        if duration_minutes is not None:
            if duration_minutes < 1:
                raise DRFValidationError({"duration_minutes": "Duration must be at least 1 minute."})
            assessment.duration_minutes = duration_minutes
        if total_points is not None:
            assessment.total_points = total_points
        if negative_marking_enabled is not None:
            assessment.negative_marking_enabled = negative_marking_enabled
        if attempt_limit is not None:
            if attempt_limit < 1:
                raise DRFValidationError({"attempt_limit": "Attempt limit must be at least 1."})
            assessment.attempt_limit = attempt_limit
        if randomize_questions is not None:
            assessment.randomize_questions = randomize_questions
        if randomize_options is not None:
            assessment.randomize_options = randomize_options
        if passing_percentage is not None:
            assessment.passing_percentage = passing_percentage
        if result_visibility is not None:
            assessment.result_visibility = result_visibility

        if assessment.end_datetime <= assessment.start_datetime:
            raise DRFValidationError({"end_datetime": "End datetime must be strictly after start datetime."})

        assessment.save()

        # Atomic draft audience persistence if audience is supplied
        if target_section_ids is not None or target_student_ids is not None:
            sec_ids = target_section_ids if target_section_ids is not None else list(assessment.target_sections.values_list('id', flat=True))
            stu_ids = target_student_ids if target_student_ids is not None else list(assessment.target_students.values_list('id', flat=True))
            AssessmentAudienceService.configure_audience(
                assessment=assessment,
                section_ids=sec_ids,
                student_ids=stu_ids,
                actor=actor,
                request=request
            )

        AuditService.log(
            action="ASSESSMENT_UPDATED",
            actor=actor,
            target_type="Assessment",
            target_id=str(assessment.id),
            metadata={"title": assessment.title},
            request=request
        )
        return assessment

    @classmethod
    @transaction.atomic
    def add_question(
        cls,
        assessment: Assessment,
        question_version: QuestionVersion,
        actor: User,
        order: Optional[int] = None,
        points: Optional[int] = None,
        negative_marking_enabled: bool = False,
        negative_points: int = 0,
        request=None
    ) -> AssessmentQuestion:
        if assessment.status in [AssessmentStatus.PUBLISHED, AssessmentStatus.ARCHIVED]:
            raise PermissionDenied("Cannot add questions to a published or archived assessment.")

        if question_version.status != VersionStatus.PUBLISHED:
            raise DRFValidationError(
                {"question_version": "Only PUBLISHED QuestionVersion records can be linked to an assessment."}
            )

        if order is None:
            max_order = assessment.assessment_questions.aggregate(models.Max('order'))['order__max'] or 0
            order = max_order + 1

        pts = points if points is not None else question_version.points
        if pts < 1:
            raise DRFValidationError({"points": "Question points must be at least 1."})

        if negative_marking_enabled and negative_points > pts:
            raise DRFValidationError({"negative_points": "Negative points penalty cannot exceed question points."})

        if assessment.assessment_questions.filter(question_version=question_version).exists():
            raise DRFValidationError({"question_version": "This question is already linked to this assessment."})

        aq = AssessmentQuestion.objects.create(
            assessment=assessment,
            question_version=question_version,
            order=order,
            points=pts,
            negative_marking_enabled=negative_marking_enabled,
            negative_points=negative_points
        )

        AuditService.log(
            action="QUESTION_ADDED_TO_ASSESSMENT",
            actor=actor,
            target_type="AssessmentQuestion",
            target_id=str(aq.id),
            metadata={
                "assessment_id": str(assessment.id),
                "question_version_id": str(question_version.id),
                "points": pts
            },
            request=request
        )
        return aq

    @classmethod
    @transaction.atomic
    def remove_question(cls, assessment: Assessment, question_version_id: str, actor: User, request=None) -> None:
        if assessment.status in [AssessmentStatus.PUBLISHED, AssessmentStatus.ARCHIVED]:
            raise PermissionDenied("Cannot remove questions from a published or archived assessment.")

        aq = assessment.assessment_questions.filter(question_version_id=question_version_id).first()
        if not aq:
            raise DRFValidationError({"question": "Question is not part of this assessment."})

        aq.delete()

        # Re-index remaining questions
        for idx, remaining in enumerate(assessment.assessment_questions.order_by('order'), start=1):
            if remaining.order != idx:
                remaining.order = idx
                remaining.save(update_fields=['order'])

        AuditService.log(
            action="QUESTION_REMOVED_FROM_ASSESSMENT",
            actor=actor,
            target_type="Assessment",
            target_id=str(assessment.id),
            metadata={"question_version_id": question_version_id},
            request=request
        )

    @classmethod
    @transaction.atomic
    def publish_assessment(
        cls,
        assessment: Assessment,
        actor: User,
        request=None,
        enforce_audience: bool = False
    ) -> Assessment:
        # 1. Lock Assessment row to prevent concurrent publish or audience mutation races
        assessment = Assessment.objects.select_for_update().get(id=assessment.id)

        if assessment.status != AssessmentStatus.DRAFT:
            raise DRFValidationError({"status": f"Only DRAFT assessments can be published. Current status: {assessment.status}"})

        qs = list(assessment.assessment_questions.all())
        if not qs:
            raise DRFValidationError({"questions": "Cannot publish an assessment with zero questions."})

        # Points Invariant SUM(AssessmentQuestion.points) == Assessment.total_points
        sum_points = sum(q.points for q in qs)
        if sum_points != assessment.total_points:
            raise DRFValidationError({
                "total_points": f"Assessment total_points ({assessment.total_points}) must exactly match the sum of question points ({sum_points})."
            })

        # 2. Fresh authoritative audience resolution inside this transaction
        resolved_audience = AssessmentAudienceService.resolve_audience(assessment)
        eligible_student_ids = resolved_audience["eligible_student_ids"]
        has_existing_assignments = assessment.assignments.filter(status=AssignmentStatus.ASSIGNED).exists()

        # 3. Zero-audience publish blocking
        if enforce_audience or (request is not None):
            if len(eligible_student_ids) == 0 and not has_existing_assignments:
                raise DRFValidationError({
                    "audience": "Select at least one student or section before publishing."
                })

        # 4. Atomic authoritative assignment creation
        if eligible_student_ids:
            cls.assign_students(
                assessment=assessment,
                student_ids=eligible_student_ids,
                actor=actor,
                request=request
            )

        # 5. Generate frozen snapshot
        AssessmentSnapshotService.create_snapshot(assessment=assessment, actor=actor, request=request)

        # 6. Transition status
        assessment.status = AssessmentStatus.PUBLISHED
        assessment.published_at = timezone.now()
        assessment.save()

        AuditService.log(
            action="ASSESSMENT_PUBLISHED",
            actor=actor,
            target_type="Assessment",
            target_id=str(assessment.id),
            metadata={
                "title": assessment.title,
                "total_points": assessment.total_points,
                "question_count": len(qs),
                "audience_assigned_count": len(eligible_student_ids)
            },
            request=request
        )
        return assessment

    @classmethod
    @transaction.atomic
    def archive_assessment(cls, assessment: Assessment, actor: User, request=None) -> Assessment:
        if assessment.status == AssessmentStatus.ARCHIVED:
            return assessment

        assessment.status = AssessmentStatus.ARCHIVED
        assessment.save()

        AuditService.log(
            action="ASSESSMENT_ARCHIVED",
            actor=actor,
            target_type="Assessment",
            target_id=str(assessment.id),
            metadata={"title": assessment.title},
            request=request
        )
        return assessment

    # --- Assessment Assignment Management ---

    @classmethod
    @transaction.atomic
    def assign_students(
        cls,
        assessment: Assessment,
        student_ids: List[str],
        actor: User,
        request=None,
        sync_draft_target: bool = False
    ) -> List[AssessmentAssignment]:
        if not student_ids:
            raise DRFValidationError({"student_ids": "At least one student ID is required."})

        # 1. Resolve & validate all student IDs BEFORE acquiring locks or mutating state
        resolved_students = resolve_student_users(student_ids, active_only=True)
        if not resolved_students:
            raise DRFValidationError({
                "error": "INVALID_STUDENT_IDS",
                "invalid_ids": student_ids,
                "detail": "None of the requested student IDs could be resolved to active students."
            })

        # 2. Lock the assessment row to serialize concurrent assignment/publish races
        assessment = Assessment.objects.select_for_update().get(id=assessment.id)

        # 3. Sort students deterministically by User.id to establish consistent lock ordering and avoid deadlocks
        sorted_students = sorted(resolved_students, key=lambda u: str(u.id))

        assignment_map = {}
        for st in sorted_students:
            assignment, created = AssessmentAssignment.objects.select_for_update().get_or_create(
                assessment=assessment,
                student=st,
                defaults={"assigned_by": actor, "status": AssignmentStatus.ASSIGNED}
            )
            if not created and assignment.status == AssignmentStatus.REVOKED:
                assignment.status = AssignmentStatus.ASSIGNED
                assignment.assigned_by = actor
                assignment.save(update_fields=['status', 'assigned_by', 'updated_at'])

            assignment_map[st.id] = assignment
            AuditService.log(
                action="ASSESSMENT_ASSIGNMENT_CREATED",
                actor=actor,
                target_type="AssessmentAssignment",
                target_id=str(assignment.id),
                metadata={"assessment_id": str(assessment.id), "student_id": str(st.id), "student_email": st.email},
                request=request
            )

        # In DRAFT status, synchronize explicit targeting intent only if requested (not during publish)
        if sync_draft_target and assessment.status == AssessmentStatus.DRAFT:
            assessment.target_students.add(*resolved_students)

        # Return assignments in caller's requested input order
        return [assignment_map[st.id] for st in resolved_students if st.id in assignment_map]

    @classmethod
    @transaction.atomic
    def revoke_assignment(cls, assessment: Assessment, student_id: str, actor: User, request=None) -> AssessmentAssignment:
        student_id_str = str(student_id).strip()
        # 1. Resolve to canonical User before acquiring locks
        try:
            resolved = resolve_student_users([student_id_str], active_only=False)
            student_user = resolved[0] if resolved else None
        except DRFValidationError:
            student_user = None

        if not student_user:
            raise DRFValidationError({"student": "Assignment does not exist for this student."})

        # 2. Lock assessment row first to adhere to uniform lock hierarchy
        assessment = Assessment.objects.select_for_update().get(id=assessment.id)

        # 3. Lock specific assignment row
        assignment = AssessmentAssignment.objects.select_for_update().filter(
            assessment=assessment,
            student=student_user
        ).first()
        if not assignment:
            raise DRFValidationError({"student": "Assignment does not exist for this student."})

        assignment.status = AssignmentStatus.REVOKED
        assignment.save(update_fields=['status', 'updated_at'])

        if assessment.status == AssessmentStatus.DRAFT:
            assessment.target_students.remove(student_user)

        AuditService.log(
            action="ASSESSMENT_ASSIGNMENT_REVOKED",
            actor=actor,
            target_type="AssessmentAssignment",
            target_id=str(assignment.id),
            metadata={"assessment_id": str(assessment.id), "student_id": str(student_user.id)},
            request=request
        )
        return assignment


class AttemptService:
    """
    Authoritative domain service for student test attempts, answer persistence, and submissions.
    """
    @classmethod
    @transaction.atomic
    def start_attempt(cls, student: User, assessment_id: str, actor: User, request=None) -> Tuple[TestAttempt, bool]:
        """
        Starts or resumes a student test attempt.
        Concurrency-safe via select_for_update and attempt_limit verification.
        """
        # Lock Assessment
        assessment = Assessment.objects.select_for_update().filter(id=assessment_id).first()
        if not assessment:
            raise DRFValidationError({"assessment": "Assessment not found."})

        if assessment.status != AssessmentStatus.PUBLISHED:
            raise DRFValidationError({"assessment": "Assessment is not currently published."})

        # Verify Student Assignment
        assignment = AssessmentAssignment.objects.filter(
            assessment=assessment,
            student=student,
            status=AssignmentStatus.ASSIGNED
        ).first()
        if not assignment:
            raise PermissionDenied("You are not assigned to take this assessment.")

        now = timezone.now()

        # Scheduling checks
        if now < assessment.start_datetime:
            raise DRFValidationError({
                "schedule": "START_REJECTED_TOO_EARLY: Assessment has not started yet.",
                "start_datetime": assessment.start_datetime.isoformat()
            })
        if now >= assessment.end_datetime:
            raise DRFValidationError({
                "schedule": "START_REJECTED_TOO_LATE: Assessment deadline has passed.",
                "end_datetime": assessment.end_datetime.isoformat()
            })

        # Check existing attempts
        existing_attempts = list(
            TestAttempt.objects.select_for_update().filter(assessment=assessment, student=student).order_by('attempt_number')
        )

        # If there is already an active IN_PROGRESS attempt, resume it
        for att in existing_attempts:
            if att.status == AttemptStatus.IN_PROGRESS:
                # Check if expired
                if AttemptTimerService.check_and_expire_attempt_if_needed(att):
                    continue
                return att, False

        # Check attempt limit
        if len(existing_attempts) >= assessment.attempt_limit:
            raise DRFValidationError({"attempt_limit": f"Attempt limit reached ({assessment.attempt_limit} max)."})

        snapshot = assessment.snapshot
        attempt_number = len(existing_attempts) + 1

        # Authoritative Seed & Randomization
        seed = RandomizationService.generate_seed()
        raw_questions = list(snapshot.snapshot_questions.order_by('order'))
        raw_q_ids = [sq.snapshot_question_id for sq in raw_questions]

        question_order = RandomizationService.randomize_question_order(
            seed=seed,
            question_ids=raw_q_ids,
            randomize=assessment.randomize_questions
        )

        option_orders = {}
        for sq in raw_questions:
            if sq.question_type in ['MCQ', 'MULTI_SELECT'] and sq.type_config:
                opts = sq.type_config.get('options', [])
                option_orders[sq.snapshot_question_id] = RandomizationService.randomize_options(
                    seed=seed,
                    question_id=sq.snapshot_question_id,
                    options=opts,
                    randomize=assessment.randomize_options
                )

        started_at = now
        expires_at = AttemptTimerService.compute_expiry(
            started_at=started_at,
            duration_minutes=assessment.duration_minutes,
            assessment_end=assessment.end_datetime
        )

        attempt = TestAttempt.objects.create(
            student=student,
            assessment=assessment,
            assessment_snapshot=snapshot,
            attempt_number=attempt_number,
            status=AttemptStatus.IN_PROGRESS,
            randomization_seed=seed,
            question_order=question_order,
            option_orders=option_orders,
            started_at=started_at,
            expires_at=expires_at
        )

        # Pre-create empty AttemptAnswer records for each snapshot question
        for sq in raw_questions:
            AttemptAnswer.objects.create(
                attempt=attempt,
                snapshot_question=sq,
                question_id=sq.snapshot_question_id,
                question_type=sq.question_type,
                revision=1,
                is_answered=False
            )

        AuditService.log(
            action="ATTEMPT_STARTED",
            actor=student,
            target_type="TestAttempt",
            target_id=str(attempt.id),
            metadata={
                "assessment_id": str(assessment.id),
                "attempt_number": attempt_number,
                "expires_at": expires_at.isoformat()
            },
            request=request
        )
        return attempt, True

    @classmethod
    def save_answer(
        cls,
        student: User,
        attempt_id: str,
        snapshot_question_id: str,
        answer_data: Dict[str, Any],
        client_revision: int = 1,
        actor: Optional[User] = None,
        request=None
    ) -> Dict[str, Any]:
        """
        Idempotently persists student answer with revision protection and expiry checks.
        """
        attempt = TestAttempt.objects.filter(id=attempt_id).first()
        if not attempt:
            raise DRFValidationError({"attempt": "Test attempt not found."})

        # Authorization: Ownership
        if attempt.student != student:
            raise PermissionDenied("You cannot save answers to another student's test attempt.")

        # Timer check & auto expire (committed immediately)
        if AttemptTimerService.check_and_expire_attempt_if_needed(attempt):
            raise DRFValidationError({"timer": "Test attempt has expired. No further answers can be saved."})

        # Status check
        if attempt.status != AttemptStatus.IN_PROGRESS:
            raise DRFValidationError({"status": f"Cannot save answer to attempt in {attempt.status} status."})

        with transaction.atomic():
            attempt_locked = TestAttempt.objects.select_for_update().get(id=attempt_id)
            if attempt_locked.status != AttemptStatus.IN_PROGRESS:
                raise DRFValidationError({"status": f"Cannot save answer to attempt in {attempt_locked.status} status."})

            # Fetch Answer record
            answer = AttemptAnswer.objects.select_for_update().filter(
                attempt=attempt_locked,
                question_id=snapshot_question_id
            ).first()

        if not answer:
            raise DRFValidationError({"question_id": "Question is not part of this attempt snapshot."})

        # Correction 8: Revision Protection
        if client_revision <= answer.revision and answer.is_answered:
            # Stale request received -> reject stale overwrite, return authoritative revision
            return {
                "status": "STALE_REVISION",
                "message": "Update ignored due to stale revision.",
                "server_revision": answer.revision,
                "is_answered": answer.is_answered,
                "last_saved_at": answer.last_saved_at.isoformat()
            }

        # Update answer fields
        is_answered = False
        if 'selected_options' in answer_data:
            answer.selected_options = answer_data['selected_options']
            if answer_data['selected_options']:
                is_answered = True

        if 'text_response' in answer_data:
            answer.text_response = answer_data['text_response']
            if answer_data['text_response'] and answer_data['text_response'].strip():
                is_answered = True

        if 'code_response' in answer_data:
            answer.code_response = answer_data['code_response']
            if answer_data['code_response'] and answer_data['code_response'].strip():
                is_answered = True

        if 'code_language' in answer_data:
            answer.code_language = answer_data['code_language']

        if 'sql_response' in answer_data:
            answer.sql_response = answer_data['sql_response']
            if answer_data['sql_response'] and answer_data['sql_response'].strip():
                is_answered = True

        answer.is_answered = is_answered
        answer.revision = max(client_revision, answer.revision + 1)
        answer.save()

        AuditService.log(
            action="ANSWER_SAVED",
            actor=actor or student,
            target_type="AttemptAnswer",
            target_id=str(answer.id),
            metadata={
                "attempt_id": str(attempt.id),
                "question_id": snapshot_question_id,
                "revision": answer.revision,
                "is_answered": is_answered
            },
            request=request
        )

        return {
            "status": "SAVED",
            "server_revision": answer.revision,
            "is_answered": answer.is_answered,
            "last_saved_at": answer.last_saved_at.isoformat()
        }

    @classmethod
    @transaction.atomic
    def submit_attempt(cls, student: User, attempt_id: str, actor: Optional[User] = None, request=None) -> TestAttempt:
        """
        Idempotent final submission of a test attempt.
        """
        attempt = TestAttempt.objects.select_for_update().filter(id=attempt_id).first()
        if not attempt:
            raise DRFValidationError({"attempt": "Test attempt not found."})

        # Authorization: Ownership
        if attempt.student != student:
            raise PermissionDenied("You cannot submit another student's test attempt.")

        # Idempotent return if already submitted
        if attempt.status == AttemptStatus.SUBMITTED:
            return attempt

        # If already expired or cancelled
        if attempt.status in [AttemptStatus.EXPIRED, AttemptStatus.CANCELLED]:
            return attempt

        attempt.status = AttemptStatus.SUBMITTED
        attempt.submitted_at = timezone.now()
        attempt.save(update_fields=['status', 'submitted_at', 'updated_at'])

        AuditService.log(
            action="ATTEMPT_SUBMITTED",
            actor=actor or student,
            target_type="TestAttempt",
            target_id=str(attempt.id),
            metadata={
                "assessment_id": str(attempt.assessment_id),
                "attempt_number": attempt.attempt_number,
                "submitted_at": attempt.submitted_at.isoformat()
            },
            request=request
        )

        try:
            from apps.results.tasks import finalize_assessment_result_task
            transaction.on_commit(lambda: finalize_assessment_result_task.delay(str(attempt.id)))
        except Exception:
            pass

        return attempt


def _sanitize_formula_injection(value: Any) -> Any:
    """
    Neutralizes CSV / Spreadsheet Formula Injection (CWE-1236).
    If a string starts with =, +, -, @, \\t, or \\r, prepends a single quote (').
    """
    if isinstance(value, str) and value:
        if value[0] in ('=', '+', '-', '@', '\t', '\r'):
            return f"'{value}"
    return value


class AssessmentAttendanceService:
    """
    Authoritative derived attendance service for CODEGUARD assessments.
    Consumes AssessmentAssignment, TestAttempt, and StudentProfile.section.
    No redundant attendance database table is created.
    """

    @classmethod
    def get_attendance_data(
        cls,
        assessment: Assessment,
        filters: Optional[Dict[str, Any]] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Dict[str, Any]:
        filters = filters or {}
        section_filter = filters.get('section_id')
        attendance_filter = filters.get('attendance_status')
        attempt_filter = filters.get('attempt_status')
        search = (filters.get('search') or '').strip().lower()

        # Authoritative query for active assignments
        assignments_qs = AssessmentAssignment.objects.filter(
            assessment=assessment,
            status=AssignmentStatus.ASSIGNED
        ).select_related(
            'student',
            'student__student_profile',
            'student__student_profile__section'
        ).prefetch_related(
            models.Prefetch(
                'student__test_attempts',
                queryset=TestAttempt.objects.filter(assessment=assessment).order_by('-attempt_number', '-created_at')
            )
        ).order_by('student__student_profile__roll_number', 'student__email')

        student_records = []
        for a in assignments_qs:
            student = a.student
            profile = getattr(student, 'student_profile', None)
            section = getattr(profile, 'section', None)
            section_name = section.name if section else 'Unassigned'
            section_id = str(section.id) if section else None

            attempts = list(student.test_attempts.all())

            # Primary attendance rule:
            # If any attempt has started_at is not None -> ATTENDED, else NOT_ATTENDED
            has_started = any(att.started_at is not None for att in attempts)
            attendance_status = 'ATTENDED' if has_started else 'NOT_ATTENDED'

            latest_att = attempts[0] if attempts else None
            attempt_status = latest_att.status if latest_att else AttemptStatus.NOT_STARTED

            started_at_str = latest_att.started_at.isoformat() if (latest_att and latest_att.started_at) else None
            submitted_at_str = latest_att.submitted_at.isoformat() if (latest_att and latest_att.submitted_at) else None

            duration_seconds = None
            if latest_att and latest_att.started_at and latest_att.submitted_at:
                duration_seconds = max(0, int((latest_att.submitted_at - latest_att.started_at).total_seconds()))

            record = {
                "student_id": str(student.id),
                "student_name": getattr(student, 'display_name', '') or student.email,
                "email": student.email,
                "roll_number": getattr(profile, 'roll_number', '') if profile else '',
                "euid": getattr(profile, 'euid', '') if profile else '',
                "section": section_name,
                "section_id": section_id,
                "assignment_status": a.status,
                "attendance_status": attendance_status,
                "attempt_status": attempt_status,
                "started_at": started_at_str,
                "submitted_at": submitted_at_str,
                "duration_seconds": duration_seconds,
                "attempts_count": len(attempts),
            }
            student_records.append(record)

        # Apply filtering
        filtered_records = []
        for r in student_records:
            if section_filter and section_filter != 'ALL':
                if section_filter == 'unassigned':
                    if r['section_id'] is not None:
                        continue
                elif r['section_id'] != str(section_filter):
                    continue

            if attendance_filter and attendance_filter != 'ALL':
                if r['attendance_status'] != attendance_filter:
                    continue

            if attempt_filter and attempt_filter != 'ALL':
                if r['attempt_status'] != attempt_filter:
                    continue

            if search:
                target_str = f"{r['student_name']} {r['email']} {r['roll_number']} {r['euid']}".lower()
                if search not in target_str:
                    continue

            filtered_records.append(r)

        # Compute summary metrics on complete FILTERED dataset before pagination
        total_assigned = len(filtered_records)
        total_attended = sum(1 for r in filtered_records if r['attendance_status'] == 'ATTENDED')
        total_not_attended = total_assigned - total_attended
        total_in_progress = sum(1 for r in filtered_records if r['attempt_status'] == AttemptStatus.IN_PROGRESS)
        total_submitted = sum(1 for r in filtered_records if r['attempt_status'] in [AttemptStatus.SUBMITTED, AttemptStatus.EXPIRED])
        attendance_percentage = round((total_attended / total_assigned * 100), 1) if total_assigned > 0 else 0.0
        is_pre_exam = assessment.start_datetime > timezone.now()

        # Section Breakdown
        from collections import OrderedDict
        section_groups = OrderedDict()
        for r in filtered_records:
            sec_key = r['section_id'] or 'unassigned'
            if sec_key not in section_groups:
                section_groups[sec_key] = {
                    "section_id": r['section_id'],
                    "section_name": r['section'],
                    "assigned": 0,
                    "attended": 0,
                    "not_attended": 0,
                    "in_progress": 0,
                    "submitted": 0,
                    "attendance_percentage": 0.0,
                }
            grp = section_groups[sec_key]
            grp["assigned"] += 1
            if r['attendance_status'] == 'ATTENDED':
                grp["attended"] += 1
            else:
                grp["not_attended"] += 1
            if r['attempt_status'] == AttemptStatus.IN_PROGRESS:
                grp["in_progress"] += 1
            elif r['attempt_status'] in [AttemptStatus.SUBMITTED, AttemptStatus.EXPIRED]:
                grp["submitted"] += 1

        for grp in section_groups.values():
            if grp["assigned"] > 0:
                grp["attendance_percentage"] = round((grp["attended"] / grp["assigned"] * 100), 1)

        # Pagination for table rows
        try:
            p = max(1, int(page))
        except (TypeError, ValueError):
            p = 1
        try:
            ps = max(1, min(100, int(page_size)))
        except (TypeError, ValueError):
            ps = 20

        start_idx = (p - 1) * ps
        end_idx = start_idx + ps
        paginated_rows = filtered_records[start_idx:end_idx]

        return {
            "assessment": {
                "id": str(assessment.id),
                "title": assessment.title,
                "status": assessment.status,
                "start_datetime": assessment.start_datetime.isoformat(),
                "end_datetime": assessment.end_datetime.isoformat(),
                "duration_minutes": assessment.duration_minutes,
                "is_pre_exam": is_pre_exam,
            },
            "summary": {
                "total_assigned": total_assigned,
                "total_attended": total_attended,
                "total_not_attended": total_not_attended,
                "total_in_progress": total_in_progress,
                "total_submitted": total_submitted,
                "attendance_percentage": attendance_percentage,
                "is_pre_exam": is_pre_exam,
            },
            "sections": list(section_groups.values()),
            "results": paginated_rows,
            "total_count": total_assigned,
            "page": p,
            "page_size": ps,
        }

    @classmethod
    def export_attendance_xlsx(cls, assessment: Assessment, filters: Optional[Dict[str, Any]] = None) -> io.BytesIO:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        data = cls.get_attendance_data(assessment, filters=filters, page=1, page_size=100000)
        summary = data['summary']
        records = data['results']
        sections = data['sections']

        wb = openpyxl.Workbook()
        ws_roster = wb.active
        ws_roster.title = "Attendance Roster"

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

        headers = [
            "Student Name", "Roll Number", "Email", "Section",
            "Assignment Status", "Attendance Status", "Attempt Status",
            "Started At", "Submitted At", "Time Used (s)"
        ]
        ws_roster.append([_sanitize_formula_injection(h) for h in headers])
        for col_num in range(1, len(headers) + 1):
            cell = ws_roster.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for r in records:
            ws_roster.append([
                _sanitize_formula_injection(r['student_name']),
                _sanitize_formula_injection(r['roll_number']),
                _sanitize_formula_injection(r['email']),
                _sanitize_formula_injection(r['section']),
                _sanitize_formula_injection(r['assignment_status']),
                _sanitize_formula_injection(r['attendance_status']),
                _sanitize_formula_injection(r['attempt_status']),
                r['started_at'] or '',
                r['submitted_at'] or '',
                r['duration_seconds'] if r['duration_seconds'] is not None else ''
            ])

        ws_sections = wb.create_sheet(title="Section Summary")
        sec_headers = ["Section", "Assigned", "Attended", "Not Attended", "In Progress", "Submitted", "Attendance %"]
        ws_sections.append([_sanitize_formula_injection(h) for h in sec_headers])
        for col_num in range(1, len(sec_headers) + 1):
            cell = ws_sections.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for sec in sections:
            ws_sections.append([
                _sanitize_formula_injection(sec['section_name']),
                sec['assigned'],
                sec['attended'],
                sec['not_attended'],
                sec['in_progress'],
                sec['submitted'],
                sec['attendance_percentage']
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @classmethod
    def export_attendance_pdf(cls, assessment: Assessment, filters: Optional[Dict[str, Any]] = None) -> io.BytesIO:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        data = cls.get_attendance_data(assessment, filters=filters, page=1, page_size=100000)
        summary = data['summary']
        records = data['results']
        sections = data['sections']

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#0F172A'),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#475569'),
            spaceAfter=14
        )

        story.append(Paragraph("CODEGUARD Assessment Attendance Report", title_style))
        pre_exam_tag = " (Pre-Exam / Scheduled)" if summary['is_pre_exam'] else ""
        story.append(Paragraph(f"Assessment: {assessment.title}{pre_exam_tag}", subtitle_style))

        meta_data = [
            ["Total Assigned:", str(summary['total_assigned']), "Attended:", str(summary['total_attended'])],
            ["Not Attended:", str(summary['total_not_attended']), "Attendance Rate:", f"{summary['attendance_percentage']}%"],
            ["In Progress:", str(summary['total_in_progress']), "Submitted:", str(summary['total_submitted'])],
        ]
        t_meta = Table(meta_data, colWidths=[110, 160, 110, 160])
        t_meta.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0F172A')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 14))

        if sections:
            story.append(Paragraph("Section Breakdown", styles['Heading3']))
            sec_table_data = [["Section", "Assigned", "Attended", "Not Attended", "In Progress", "Submitted", "Attendance %"]]
            for sec in sections:
                sec_table_data.append([
                    sec['section_name'],
                    str(sec['assigned']),
                    str(sec['attended']),
                    str(sec['not_attended']),
                    str(sec['in_progress']),
                    str(sec['submitted']),
                    f"{sec['attendance_percentage']}%"
                ])
            t_sec = Table(sec_table_data, colWidths=[130, 65, 65, 75, 70, 65, 70])
            t_sec.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ]))
            story.append(t_sec)
            story.append(Spacer(1, 14))

        story.append(Paragraph("Student Attendance Roster", styles['Heading3']))
        roster_table_data = [["Student Name", "Roll No", "Section", "Attendance", "Attempt", "Started At"]]
        for r in records[:100]:
            roster_table_data.append([
                r['student_name'][:24],
                r['roll_number'] or 'N/A',
                r['section'][:16],
                r['attendance_status'],
                r['attempt_status'],
                r['started_at'][:19].replace('T', ' ') if r['started_at'] else 'N/A'
            ])
        t_roster = Table(roster_table_data, colWidths=[120, 75, 85, 75, 85, 100])
        t_roster.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ]))
        story.append(t_roster)

        doc.build(story)
        output.seek(0)
        return output

