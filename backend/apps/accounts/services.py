import io
import re
import csv
import logging
import secrets
import string
import uuid
from typing import List, Dict, Tuple, Any, Optional
import openpyxl
from django.db import transaction, IntegrityError
from django.core.validators import validate_email
from django.core.exceptions import ValidationError as DjangoValidationError
from django.contrib.auth.password_validation import validate_password
from django.contrib.sessions.models import Session
from rest_framework.exceptions import ValidationError as DRFValidationError

from .models import User, StudentProfile, AuditLog, Role, AdminSequence, Section

logger = logging.getLogger(__name__)

class AdminIdService:
    """
    Authoritative Admin ID generation service.
    Primary Admin ID: EUAD-GAURAV-099
    Secondary Admin IDs: CG-ADM-000002, CG-ADM-000003, ...
    Note: CG-ADM-000001 is permanently retired and will never be assigned.
    Uses select_for_update() on AdminSequence to prevent race conditions during concurrent creation.
    """
    @classmethod
    def generate_next_admin_id(cls) -> str:
        with transaction.atomic():
            seq, _ = AdminSequence.objects.select_for_update().get_or_create(
                id=uuid.UUID('00000000-0000-0000-0000-000000000001'),
                defaults={'last_sequence': 2}
            )
            # Never start below 2 (CG-ADM-000001 is retired)
            current = max(2, seq.last_sequence)
            while True:
                candidate = f"CG-ADM-{current:06d}"
                if candidate == 'CG-ADM-000001':
                    current += 1
                    continue
                if not User.objects.filter(admin_id=candidate).exists():
                    break
                current += 1
            seq.last_sequence = current + 1
            seq.save(update_fields=['last_sequence', 'updated_at'])
            return candidate

class EUIDService:
    """
    Deterministic, collision-safe Exam Unique ID (EUID) generation service.
    Generates standardized EUIDs derived from student academic roll numbers.
    Format: CG-{NORMALIZED_ROLL_NUMBER} (e.g. CG-BETN1AI25099)
    Strictly rejects collisions without generating numbered suffixes.
    """
    @staticmethod
    def normalize_roll_number(roll_number: str) -> str:
        if not roll_number:
            raise DRFValidationError("Roll number cannot be empty.")
        clean = roll_number.strip().upper()
        # Remove unwanted punctuation while preserving alphanumeric and hyphens/underscores
        normalized = re.sub(r'[^A-Z0-9_-]', '', clean)
        if not normalized:
            raise DRFValidationError("Roll number contains no valid alphanumeric characters.")
        return normalized

    @classmethod
    def generate_euid(cls, roll_number: str) -> str:
        normalized_roll = cls.normalize_roll_number(roll_number)
        return f"CG-{normalized_roll}"

    @classmethod
    def validate_unique_euid(cls, roll_number: str, exclude_user_id: Optional[str] = None) -> str:
        euid = cls.generate_euid(roll_number)
        qs = StudentProfile.objects.filter(euid=euid)
        if exclude_user_id:
            qs = qs.exclude(user__id=exclude_user_id)
        if qs.exists():
            raise DRFValidationError({"euid": f"A student with EUID '{euid}' already exists in the system."})
        return euid


class AuditService:
    """
    Immutable audit logging service recording administrative and security actions.
    """
    @staticmethod
    def extract_ip_address(request) -> Optional[str]:
        if not request:
            return None
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR')

    @classmethod
    def log(
        cls,
        action: str,
        actor: Optional[User] = None,
        target_type: Optional[str] = None,
        target_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
        request=None
    ) -> AuditLog:
        ip = cls.extract_ip_address(request) if request else None
        safe_meta = metadata.copy() if metadata else {}
        
        # Ensure password or credential tokens never leak into audit metadata
        for sensitive_key in ['password', 'password_hash', 'current_password', 'new_password', 'token']:
            safe_meta.pop(sensitive_key, None)

        return AuditLog.objects.create(
            actor=actor,
            action=action,
            target_type=target_type,
            target_id=str(target_id) if target_id else None,
            metadata=safe_meta,
            ip_address=ip
        )


class SectionService:
    """
    Authoritative domain service managing Academic Section entities, normalization,
    race-safe uniqueness, student counts, and safe historical deactivation.
    """
    @staticmethod
    def normalize_code(code: str) -> str:
        if not code or not str(code).strip():
            raise DRFValidationError({"code": "Section code cannot be empty."})
        clean = str(code).strip().upper()
        normalized = re.sub(r'[^A-Z0-9_-]', '', clean)
        if not normalized:
            raise DRFValidationError({"code": "Section code contains no valid characters."})
        return normalized

    @classmethod
    def create_section(
        cls,
        code: str,
        name: str,
        is_active: bool = True,
        actor: Optional[User] = None,
        request=None
    ) -> Section:
        norm_code = cls.normalize_code(code)
        clean_name = str(name).strip() if name else ""
        if not clean_name:
            raise DRFValidationError({"name": "Section name cannot be empty."})

        if Section.objects.filter(code=norm_code).exists():
            raise DRFValidationError({"code": f"A section with code '{norm_code}' already exists."})

        try:
            with transaction.atomic():
                section = Section.objects.create(
                    code=norm_code,
                    name=clean_name,
                    is_active=is_active
                )
                AuditService.log(
                    action="SECTION_CREATED",
                    actor=actor,
                    target_type="Section",
                    target_id=str(section.id),
                    metadata={
                        "code": section.code,
                        "name": section.name,
                        "is_active": section.is_active
                    },
                    request=request
                )
                return section
        except IntegrityError:
            raise DRFValidationError({"code": f"A section with code '{norm_code}' already exists."})

    @classmethod
    def update_section(
        cls,
        section: Section,
        name: Optional[str] = None,
        is_active: Optional[bool] = None,
        actor: Optional[User] = None,
        request=None
    ) -> Section:
        update_fields = ['updated_at']
        old_active = section.is_active

        if name is not None:
            clean_name = str(name).strip()
            if not clean_name:
                raise DRFValidationError({"name": "Section name cannot be empty."})
            section.name = clean_name
            update_fields.append('name')

        if is_active is not None and is_active != old_active:
            section.is_active = is_active
            update_fields.append('is_active')

        section.save(update_fields=update_fields)

        action = "SECTION_UPDATED"
        if is_active is not None and is_active != old_active:
            action = "SECTION_ACTIVATED" if is_active else "SECTION_DEACTIVATED"

        AuditService.log(
            action=action,
            actor=actor,
            target_type="Section",
            target_id=str(section.id),
            metadata={
                "code": section.code,
                "name": section.name,
                "is_active": section.is_active
            },
            request=request
        )
        return section

    @classmethod
    def delete_or_deactivate_section(
        cls,
        section: Section,
        actor: Optional[User] = None,
        request=None
    ) -> None:
        """
        Safe deletion check: Rejects deletion if students or assessment targeting references exist,
        recommending deactivation instead.
        """
        has_students = section.students.exists()
        has_assessments = section.targeted_assessments.exists()

        if has_students or has_assessments:
            raise DRFValidationError({
                "detail": "This section has existing dependencies. Deactivate it instead of deleting it."
            })

        sec_id = str(section.id)
        sec_code = section.code
        with transaction.atomic():
            section.delete()
            AuditService.log(
                action="SECTION_DELETED",
                actor=actor,
                target_type="Section",
                target_id=sec_id,
                metadata={"code": sec_code},
                request=request
            )

    @classmethod
    def get_sections_with_counts(cls, active_only: bool = False):
        from django.db.models import Count, Q
        qs = Section.objects.annotate(
            student_count=Count('students', filter=Q(students__user__is_active=True))
        ).order_by('code')
        if active_only:
            qs = qs.filter(is_active=True)
        return qs


class StudentService:
    """
    Domain service orchestrating student account lifecycle, verification, and creation.
    """
    @staticmethod
    def create_student(
        email: str,
        roll_number: str,
        section: Optional[Section] = None,
        actor: Optional[User] = None,
        request=None
    ) -> Tuple[User, StudentProfile]:
        clean_email = email.strip().lower()
        clean_roll = EUIDService.normalize_roll_number(roll_number)

        if section is not None and not section.is_active:
            raise DRFValidationError({"section": f"Cannot assign inactive section '{section.code}'."})

        try:
            validate_email(clean_email)
        except DjangoValidationError:
            raise DRFValidationError({"email": "Enter a valid email address."})

        if User.objects.filter(email=clean_email).exists():
            raise DRFValidationError({"email": "A user with this email address already exists."})

        if StudentProfile.objects.filter(roll_number=clean_roll).exists():
            raise DRFValidationError({"roll_number": "A student with this roll number already exists."})

        euid = EUIDService.validate_unique_euid(clean_roll)

        try:
            with transaction.atomic():
                # Initial password is the exact student roll number (hashed securely)
                user = User.objects.create_user(
                    email=clean_email,
                    password=clean_roll,
                    role=Role.STUDENT,
                    is_active=True,
                    first_login_required=True
                )

                profile = StudentProfile.objects.create(
                    user=user,
                    section=section,
                    roll_number=clean_roll,
                    euid=euid,
                    first_login_required=True
                )

                AuditService.log(
                    action="STUDENT_CREATED",
                    actor=actor,
                    target_type="StudentProfile",
                    target_id=str(profile.id),
                    metadata={
                        "email": user.email,
                        "roll_number": profile.roll_number,
                        "euid": profile.euid,
                        "section": section.code if section else None
                    },
                    request=request
                )
        except IntegrityError:
            raise DRFValidationError({"roll_number": "A student with this roll number or EUID already exists."})

        return user, profile

    @staticmethod
    def change_student_section(
        student_profile: StudentProfile,
        new_section: Optional[Section],
        actor: Optional[User] = None,
        request=None
    ) -> StudentProfile:
        """
        Classify/reclassify a student into a new academic section.
        Section is purely a classification mechanism:
        Existing assessment assignments are NEVER modified, revoked, or automatically granted.
        """
        if hasattr(student_profile, 'student_profile'):
            student_profile = student_profile.student_profile

        if new_section is not None and not new_section.is_active:
            raise DRFValidationError({"section": f"Cannot assign inactive section '{new_section.code}'."})

        old_section_code = student_profile.section.code if student_profile.section else None
        new_section_code = new_section.code if new_section else None

        if old_section_code == new_section_code:
            return student_profile

        student_profile.section = new_section
        student_profile.save(update_fields=['section', 'updated_at'])

        AuditService.log(
            action="STUDENT_SECTION_CHANGED",
            actor=actor,
            target_type="StudentProfile",
            target_id=str(student_profile.id),
            metadata={
                "actor_name": actor.display_name if actor else "SYSTEM",
                "target_identity": student_profile.euid,
                "roll_number": student_profile.roll_number,
                "old_section": old_section_code,
                "new_section": new_section_code
            },
            request=request
        )
        return student_profile

    @staticmethod
    def update_student(
        student_profile: StudentProfile,
        email: Optional[str] = None,
        section: Optional[Section] = None,
        update_section: bool = False,
        actor: Optional[User] = None,
        request=None
    ) -> StudentProfile:
        user = student_profile.user
        if email:
            clean_email = email.strip().lower()
            if clean_email != user.email:
                try:
                    validate_email(clean_email)
                except DjangoValidationError:
                    raise DRFValidationError({"email": "Enter a valid email address."})

                if User.objects.filter(email=clean_email).exclude(id=user.id).exists():
                    raise DRFValidationError({"email": "Email is already taken by another account."})

                with transaction.atomic():
                    user.email = clean_email
                    user.save(update_fields=['email', 'updated_at'])

                    AuditService.log(
                        action="STUDENT_UPDATED",
                        actor=actor,
                        target_type="StudentProfile",
                        target_id=str(student_profile.id),
                        metadata={"email": clean_email},
                        request=request
                    )

        if update_section:
            StudentService.change_student_section(
                student_profile=student_profile,
                new_section=section,
                actor=actor,
                request=request
            )

        return student_profile

    @staticmethod
    def set_student_status(
        student_profile: StudentProfile,
        is_active: bool,
        actor: Optional[User] = None,
        reason: str = "",
        request=None
    ) -> StudentProfile:
        user = student_profile.user
        if user.is_active == is_active:
            return student_profile

        user.is_active = is_active
        user.save(update_fields=['is_active', 'updated_at'])

        if not is_active:
            AccountSecurityService.revoke_user_sessions(user.id)

        action = "STUDENT_ENABLED" if is_active else "STUDENT_DISABLED"
        AuditService.log(
            action=action,
            actor=actor,
            target_type="StudentProfile",
            target_id=student_profile.id,
            metadata={
                "actor_name": actor.display_name if actor else "SYSTEM",
                "actor_admin_id": getattr(actor, 'admin_id', ""),
                "target_identity": student_profile.euid,
                "target_roll_number": student_profile.roll_number,
                "target_email": user.email,
                "target_role": Role.STUDENT,
                "is_active": is_active,
                "reason": reason.strip() if reason else ("Account enabled by administrator." if is_active else "Account disabled by administrator."),
                "result": "SUCCESS"
            },
            request=request
        )

        return student_profile

    @staticmethod
    def delete_student(
        student_profile: StudentProfile,
        actor: Optional[User] = None,
        request=None
    ) -> None:
        """
        Delete a student account while strictly respecting Phase 9 retention and legal holds.
        If the student has examination attempts, results, retention records, or active legal holds,
        physical deletion is rejected with a clear architectural explanation.
        """
        user = student_profile.user

        from apps.assessments.models import TestAttempt
        from apps.retention.models import LegalHold, LegalHoldStatus, RetentionRecord, ExportJob, ExportStatus
        from apps.results.models import AssessmentResult
        from django.db.models import Q

        has_attempts = TestAttempt.objects.filter(student=user).exists()
        has_results = AssessmentResult.objects.filter(student=user).exists()
        has_retention = RetentionRecord.objects.filter(attempt__student=user).exists()
        has_legal_hold = LegalHold.objects.filter(
            Q(student=user) | Q(attempt__student=user),
            status=LegalHoldStatus.ACTIVE
        ).exists()

        if has_attempts or has_results or has_retention or has_legal_hold:
            raise DRFValidationError({
                "detail": "Deletion cannot be completed yet because this account has retained examination records / an active legal hold."
            })

        has_in_flight_export = ExportJob.objects.filter(
            student=user,
            status__in=[ExportStatus.REQUESTED, ExportStatus.SNAPSHOT_PENDING, ExportStatus.GENERATING]
        ).exists()
        if has_in_flight_export:
            raise DRFValidationError({
                "detail": "Deletion cannot be completed while a DSAR data export archive is actively generating."
            })

        AccountSecurityService.revoke_user_sessions(user.id)

        AuditService.log(
            action="STUDENT_DELETED",
            actor=actor,
            target_type="StudentProfile",
            target_id=str(student_profile.id),
            metadata={
                "actor_name": actor.display_name if actor else "SYSTEM",
                "actor_admin_id": getattr(actor, 'admin_id', ""),
                "target_identity": student_profile.euid,
                "target_roll_number": student_profile.roll_number,
                "target_email": user.email,
                "target_role": Role.STUDENT,
                "reason": "Student account removed by authorized administrator",
                "result": "SUCCESS"
            },
            request=request
        )

        with transaction.atomic():
            student_profile.delete()
            user.delete()


class ImportService:
    """
    Bulk student import service supporting CSV and XLSX formats with a strict
    Validation -> Preview -> Admin Confirmation workflow.
    """
    MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024  # 5 MB
    ALLOWED_EXTENSIONS = ['.csv', '.xlsx']

    @classmethod
    def parse_file(cls, uploaded_file) -> List[Dict[str, str]]:
        filename = uploaded_file.name.lower()
        if not any(filename.endswith(ext) for ext in cls.ALLOWED_EXTENSIONS):
            raise DRFValidationError(f"Invalid file format. Supported formats: {', '.join(cls.ALLOWED_EXTENSIONS)}")

        if uploaded_file.size > cls.MAX_FILE_SIZE_BYTES:
            raise DRFValidationError(f"File size exceeds maximum allowed limit ({cls.MAX_FILE_SIZE_BYTES // (1024*1024)} MB).")

        rows: List[Dict[str, str]] = []

        if filename.endswith('.csv'):
            try:
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                header_row = next(reader, None)
                if not header_row:
                    raise DRFValidationError("The uploaded CSV file is empty.")

                roll_idx, email_idx, section_idx = cls._detect_columns(header_row)

                for line_num, row in enumerate(reader, start=2):
                    if not any(cell.strip() for cell in row):
                        continue  # Skip blank lines
                    roll = row[roll_idx].strip() if len(row) > roll_idx else ""
                    email = row[email_idx].strip() if len(row) > email_idx else ""
                    has_sec = (section_idx != -1)
                    section = row[section_idx].strip() if has_sec and len(row) > section_idx else ""
                    rows.append({
                        "row_number": line_num,
                        "roll_number": roll,
                        "email": email,
                        "section": section,
                        "has_section_column": has_sec
                    })

            except UnicodeDecodeError:
                raise DRFValidationError("Failed to decode CSV file. Please ensure it is saved in UTF-8 encoding.")
            except Exception as e:
                if isinstance(e, DRFValidationError):
                    raise e
                raise DRFValidationError(f"Error parsing CSV file: {str(e)}")

        elif filename.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                sheet = wb.active
                iter_rows = sheet.iter_rows(values_only=True)
                header_row = next(iter_rows, None)
                if not header_row:
                    raise DRFValidationError("The uploaded Excel workbook is empty.")

                header_list = [str(col).strip() if col is not None else "" for col in header_row]
                roll_idx, email_idx, section_idx = cls._detect_columns(header_list)

                for line_num, row in enumerate(iter_rows, start=2):
                    if not row or not any(str(cell).strip() for cell in row if cell is not None):
                        continue
                    roll = str(row[roll_idx]).strip() if len(row) > roll_idx and row[roll_idx] is not None else ""
                    email = str(row[email_idx]).strip() if len(row) > email_idx and row[email_idx] is not None else ""
                    has_sec = (section_idx != -1)
                    section = str(row[section_idx]).strip() if has_sec and len(row) > section_idx and row[section_idx] is not None else ""
                    rows.append({
                        "row_number": line_num,
                        "roll_number": roll,
                        "email": email,
                        "section": section,
                        "has_section_column": has_sec
                    })

            except Exception as e:
                if isinstance(e, DRFValidationError):
                    raise e
                raise DRFValidationError(f"Error parsing Excel spreadsheet: {str(e)}")

        if not rows:
            raise DRFValidationError("No student data rows found in the uploaded file.")

        return rows

    @staticmethod
    def _detect_columns(header_row: List[str]) -> Tuple[int, int, int]:
        roll_aliases = ['roll number', 'roll_number', 'roll', 'rollno', 'roll_no', 'student roll number', 'id']
        email_aliases = ['email', 'email address', 'email_address', 'student email', 'mail']
        section_aliases = ['section', 'section_code', 'section code', 'sec', 'class section', 'class_section']

        roll_idx = -1
        email_idx = -1
        section_idx = -1

        for idx, col in enumerate(header_row):
            col_clean = str(col).strip().lower()
            if col_clean in roll_aliases and roll_idx == -1:
                roll_idx = idx
            elif col_clean in email_aliases and email_idx == -1:
                email_idx = idx
            elif col_clean in section_aliases and section_idx == -1:
                section_idx = idx

        if roll_idx == -1 or email_idx == -1:
            raise DRFValidationError(
                "Required column headers not found. Expected headers: 'Roll Number' and 'Email' (optional 'Section')."
            )

        return roll_idx, email_idx, section_idx

    @classmethod
    def validate_preview(cls, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Validates parsed rows independently, checks database uniqueness, section existence/state,
        and generates an interactive preview report with actionable row errors.
        """
        seen_rolls: Dict[str, int] = {}
        seen_emails: Dict[str, int] = {}

        # Collect candidate identifiers for single-query bulk DB lookup
        candidate_rolls = set()
        candidate_emails = set()

        for r in rows:
            roll = r.get('roll_number', '').strip().upper()
            email = r.get('email', '').strip().lower()
            if roll:
                candidate_rolls.add(roll)
            if email:
                candidate_emails.add(email)

        existing_rolls = set(StudentProfile.objects.filter(roll_number__in=candidate_rolls).values_list('roll_number', flat=True))
        existing_emails = set(User.objects.filter(email__in=candidate_emails).values_list('email', flat=True))

        # Fetch all registered sections
        all_sections = {s.code: s for s in Section.objects.all()}

        preview_rows = []
        valid_count = 0
        invalid_count = 0
        duplicate_count = 0

        for r in rows:
            row_num = r.get('row_number', 0)
            raw_roll = r.get('roll_number', '').strip()
            raw_email = r.get('email', '').strip()
            raw_sec = r.get('section', '').strip()
            has_section_column = r.get('has_section_column', False)
            errors = []
            status_tag = "VALID"

            # 1. Roll Number Validation
            norm_roll = ""
            if not raw_roll:
                errors.append("Roll number is missing.")
            else:
                try:
                    norm_roll = EUIDService.normalize_roll_number(raw_roll)
                    euid_candidate = EUIDService.generate_euid(raw_roll)
                    if norm_roll in seen_rolls:
                        errors.append(f"Duplicate roll number in file (first seen at row {seen_rolls[norm_roll]}).")
                        status_tag = "DUPLICATE"
                    elif norm_roll in existing_rolls:
                        errors.append(f"Roll number already exists in system database (EUID: {euid_candidate}).")
                        status_tag = "DUPLICATE"
                    else:
                        seen_rolls[norm_roll] = row_num
                except DRFValidationError as ve:
                    errors.append(str(ve.detail[0] if isinstance(ve.detail, list) else ve.detail))

            # 2. Email Validation
            clean_email = ""
            if not raw_email:
                errors.append("Email address is missing.")
            else:
                clean_email = raw_email.lower()
                try:
                    validate_email(clean_email)
                    if clean_email in seen_emails:
                        errors.append(f"Duplicate email address in file (first seen at row {seen_emails[clean_email]}).")
                        status_tag = "DUPLICATE"
                    elif clean_email in existing_emails:
                        errors.append("Email address already exists in system database.")
                        status_tag = "DUPLICATE"
                    else:
                        seen_emails[clean_email] = row_num
                except DjangoValidationError:
                    errors.append("Invalid email address format.")

            # 3. Section Validation
            if has_section_column:
                # Modern import with Section column: section is strictly required for each row
                if not raw_sec:
                    errors.append("Section is required")
                else:
                    norm_sec = raw_sec.upper()
                    sec_obj = all_sections.get(norm_sec)
                    if not sec_obj:
                        errors.append(f"Unknown section: {raw_sec}")
                    elif not sec_obj.is_active:
                        errors.append(f"Section is inactive: {raw_sec}")
            else:
                # Legacy import without section column: allowed for compatibility
                if raw_sec:
                    norm_sec = raw_sec.upper()
                    sec_obj = all_sections.get(norm_sec)
                    if not sec_obj:
                        errors.append(f"Unknown section: {raw_sec}")
                    elif not sec_obj.is_active:
                        errors.append(f"Section is inactive: {raw_sec}")

            if errors:
                if status_tag != "DUPLICATE":
                    status_tag = "INVALID"
                if status_tag == "DUPLICATE":
                    duplicate_count += 1
                else:
                    invalid_count += 1
            else:
                valid_count += 1

            euid_preview = EUIDService.generate_euid(raw_roll) if raw_roll else ""

            preview_rows.append({
                "row_number": row_num,
                "roll_number": raw_roll,
                "email": raw_email,
                "section": raw_sec,
                "euid": euid_preview,
                "status": status_tag,
                "errors": errors
            })

        return {
            "total_rows": len(rows),
            "valid_count": valid_count,
            "invalid_count": invalid_count,
            "duplicate_count": duplicate_count,
            "rows": preview_rows
        }

    @classmethod
    def execute_import(
        cls,
        items: List[Dict[str, str]],
        actor: Optional[User] = None,
        filename: Optional[str] = None,
        request=None
    ) -> Dict[str, Any]:
        """
        Executes verified student creation for all valid rows inside an atomic transaction.
        """
        created_students = []
        failed_rows = []

        with transaction.atomic():
            for item in items:
                roll = item.get('roll_number', '').strip()
                email = item.get('email', '').strip().lower()
                sec_code = item.get('section', '').strip()

                sec_obj = None
                if sec_code:
                    sec_obj = Section.objects.filter(code=sec_code.upper()).first()

                try:
                    user, profile = StudentService.create_student(
                        email=email,
                        roll_number=roll,
                        section=sec_obj,
                        actor=actor,
                        request=request
                    )
                    created_students.append({
                        "id": str(profile.id),
                        "email": user.email,
                        "roll_number": profile.roll_number,
                        "euid": profile.euid,
                        "section": profile.section.code if profile.section else None
                    })
                except Exception as e:
                    failed_rows.append({
                        "roll_number": roll,
                        "email": email,
                        "section": sec_code,
                        "error": str(e)
                    })

            AuditService.log(
                action="STUDENT_IMPORT_COMPLETED",
                actor=actor,
                target_type="StudentProfile",
                metadata={
                    "filename": filename,
                    "total_submitted": len(items),
                    "created_count": len(created_students),
                    "failed_count": len(failed_rows)
                },
                request=request
            )

        return {
            "total_submitted": len(items),
            "created_count": len(created_students),
            "failed_count": len(failed_rows),
            "created_students": created_students,
            "failed_rows": failed_rows
        }


class AccountSecurityService:
    """
    Domain service orchestrating administrative password resets, session invalidation,
    and high-security credential lifecycle operations.
    """

    @staticmethod
    def generate_secure_temporary_password(length: int = 14) -> str:
        """
        Generates a cryptographically random, high-entropy temporary password
        meeting all Django complexity standards (upper, lower, digit, special character).
        """
        alphabet = string.ascii_letters + string.digits + "!@#$%&*+?"
        while True:
            password = ''.join(secrets.choice(alphabet) for _ in range(length))
            if (any(c.islower() for c in password)
                    and any(c.isupper() for c in password)
                    and any(c.isdigit() for c in password)
                    and any(c in "!@#$%&*+?" for c in password)):
                return password

    @classmethod
    def revoke_user_sessions(cls, user_id: Any) -> int:
        """
        Hard-revokes all active Django sessions associated with the given user ID.
        """
        revoked_count = 0
        try:
            for session in Session.objects.all():
                try:
                    data = session.get_decoded()
                    if str(data.get('_auth_user_id')) == str(user_id):
                        session.delete()
                        revoked_count += 1
                except Exception:
                    continue
        except Exception as e:
            logger.warning("Session revocation error for user %s: %s", user_id, str(e))
        return revoked_count

    @classmethod
    def reset_student_password(
        cls,
        student_profile: StudentProfile,
        temporary_password: Optional[str] = None,
        reason: str = "",
        actor: Optional[User] = None,
        request=None
    ) -> str:
        """
        Admin resets a student's password.
        1. Accepts or generates secure temporary password.
        2. Validates against Django password rules.
        3. Updates user hash and flags first_login_required.
        4. Hard-revokes active student sessions.
        5. Writes immutable AuditLog (zero sensitive credential leakage).
        6. Returns temporary password for one-time display.
        """
        user = student_profile.user
        clean_reason = reason.strip() if reason else ""
        if not clean_reason:
            raise DRFValidationError({"reason": "An administrative reason is required for password reset."})

        temp_pwd = temporary_password.strip() if temporary_password else cls.generate_secure_temporary_password()
        validate_password(temp_pwd, user=user)

        with transaction.atomic():
            user.set_password(temp_pwd)
            user.first_login_required = True
            user.save(update_fields=['password', 'first_login_required', 'updated_at'])

            student_profile.first_login_required = True
            student_profile.save(update_fields=['first_login_required', 'updated_at'])

            cls.revoke_user_sessions(user.id)

            AuditService.log(
                action="PASSWORD_RESET",
                actor=actor,
                target_type="StudentProfile",
                target_id=str(student_profile.id),
                metadata={
                    "actor_name": actor.display_name if actor else "SYSTEM",
                    "actor_admin_id": getattr(actor, 'admin_id', ""),
                    "target_identity": student_profile.euid,
                    "target_roll_number": student_profile.roll_number,
                    "target_email": user.email,
                    "target_role": Role.STUDENT,
                    "reason": clean_reason,
                    "result": "SUCCESS"
                },
                request=request
            )

        return temp_pwd

    @classmethod
    def reset_admin_password(
        cls,
        target_admin: User,
        temporary_password: Optional[str] = None,
        reason: str = "",
        actor: Optional[User] = None,
        request=None
    ) -> str:
        """
        Admin resets another administrator's password.
        1. Rejects self-reset (directing to profile security settings).
        2. Accepts or generates secure temporary password.
        3. Validates against Django password rules.
        4. Updates admin password hash and flags first_login_required.
        5. Hard-revokes active admin sessions without affecting actor's session.
        6. Writes immutable AuditLog.
        7. Returns temporary password for one-time display.
        """
        if target_admin.role != Role.ADMIN:
            raise DRFValidationError({"detail": "Target user is not an administrator."})

        if actor and str(actor.id) == str(target_admin.id):
            raise DRFValidationError({"detail": "Use Change Password in your account security settings to change your own password."})

        # If target is primary admin, only primary admin can reset
        if target_admin.is_primary_admin and actor and not actor.is_primary_admin:
            raise DRFValidationError({"detail": "Primary Administrator account cannot be reset by secondary administrators."})

        clean_reason = reason.strip() if reason else ""
        if not clean_reason:
            raise DRFValidationError({"reason": "An administrative reason is required for password reset."})

        temp_pwd = temporary_password.strip() if temporary_password else cls.generate_secure_temporary_password()
        validate_password(temp_pwd, user=target_admin)

        with transaction.atomic():
            target_admin.set_password(temp_pwd)
            target_admin.first_login_required = True
            target_admin.save(update_fields=['password', 'first_login_required', 'updated_at'])

            cls.revoke_user_sessions(target_admin.id)

            AuditService.log(
                action="PASSWORD_RESET",
                actor=actor,
                target_type="User",
                target_id=str(target_admin.id),
                metadata={
                    "actor_name": actor.display_name if actor else "SYSTEM",
                    "actor_admin_id": getattr(actor, 'admin_id', ""),
                    "target_identity": target_admin.admin_id,
                    "target_email": target_admin.email,
                    "target_role": Role.ADMIN,
                    "reason": clean_reason,
                    "result": "SUCCESS"
                },
                request=request
            )

        return temp_pwd

    @classmethod
    def update_administrator(
        cls,
        target_admin: User,
        display_name: Optional[str] = None,
        email: Optional[str] = None,
        actor: Optional[User] = None,
        request=None
    ) -> User:
        """
        Decommissioned: Administrator identity is strictly immutable and cannot be updated.
        """
        raise DRFValidationError({"detail": "Administrator identity is strictly immutable and cannot be updated."})

    @classmethod
    def delete_administrator(
        cls,
        target_admin: User,
        actor: Optional[User] = None,
        request=None
    ) -> None:
        """
        Safely delete a secondary administrator account.
        Primary Admin and self cannot be deleted.
        Checks active duties, active legal holds, and retention policies.
        Revokes sessions, logs immutable audit snapshot, and removes user.
        """
        if target_admin.role != Role.ADMIN:
            raise DRFValidationError({"detail": "Target user is not an administrator."})

        if target_admin.is_primary_admin:
            raise DRFValidationError({"detail": "The Primary Administrator account cannot be deleted."})

        if actor and str(actor.id) == str(target_admin.id):
            raise DRFValidationError({"detail": "Cannot delete your own administrator account."})

        from apps.invigilation.models import ProctorDutySession, ProctorAssignment
        from apps.retention.models import LegalHold, LegalHoldStatus, RetentionPolicy

        if ProctorDutySession.objects.filter(proctor=target_admin, is_active=True).exists():
            raise DRFValidationError({"detail": "Cannot delete administrator with active proctor duty sessions. End active sessions first."})

        if ProctorAssignment.objects.filter(proctor=target_admin, is_active=True).exists():
            raise DRFValidationError({"detail": "Cannot delete administrator assigned to active proctoring cohorts. Deactivate assignments first."})

        if LegalHold.objects.filter(placed_by=target_admin, status=LegalHoldStatus.ACTIVE).exists():
            raise DRFValidationError({"detail": "Cannot delete administrator with active placed legal holds. Release or reassign holds first."})

        if RetentionPolicy.objects.filter(created_by=target_admin, is_active=True).exists():
            raise DRFValidationError({"detail": "Cannot delete administrator with active created retention policies."})

        cls.revoke_user_sessions(target_admin.id)

        AuditService.log(
            action="ADMIN_DELETED",
            actor=actor,
            target_type="User",
            target_id=str(target_admin.id),
            metadata={
                "actor_name": actor.display_name if actor else "SYSTEM",
                "actor_admin_id": getattr(actor, 'admin_id', ""),
                "target_identity": target_admin.admin_id,
                "target_name": target_admin.display_name,
                "target_email": target_admin.email,
                "target_role": Role.ADMIN,
                "reason": "Secondary administrator account deleted by Primary Administrator",
                "result": "SUCCESS"
            },
            request=request
        )

        with transaction.atomic():
            target_admin.delete()


