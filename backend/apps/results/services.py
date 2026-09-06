import io
import os
import math
import hashlib
import logging
from decimal import Decimal
from typing import Optional, Dict, Any, List, Tuple
from datetime import timedelta

logger = logging.getLogger(__name__)

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from django.db import models, transaction
from django.conf import settings
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from rest_framework.exceptions import ValidationError as DRFValidationError, NotFound

from apps.accounts.models import User, Role
from apps.accounts.services import AuditService
from apps.assessments.models import (
    Assessment,
    AssessmentStatus,
    AssessmentSnapshot,
    AssessmentSnapshotQuestion,
    TestAttempt,
    AttemptStatus,
    AttemptAnswer,
    ResultVisibility,
)
from apps.questions.models import QuestionType
from apps.evaluator.models import CodeSubmission, SubmissionType, CodeVerdict
from apps.proctoring.models import ProctoringSession, RiskBand
from .models import (
    AssessmentResult,
    QuestionResult,
    HistoricalResultSummary,
    AssessmentAnalyticsSnapshot,
    ReportJob,
    ResultStatus,
    ReportType,
    ReportFormat,
    ReportStatus,
)


def _sanitize_formula_injection(value: Any) -> Any:
    """
    Neutralizes CSV / Spreadsheet Formula Injection (CWE-1236).
    If a string starts with =, +, -, @, \\t, or \\r, prepends a single quote (').
    """
    if isinstance(value, str) and value:
        if value[0] in ('=', '+', '-', '@', '\t', '\r'):
            return f"'{value}"
    return value


class ResultAccessPolicyService:
    """
    Enforces Phase 5 ResultVisibility and object ownership rules on result access.
    """
    @staticmethod
    def can_view_result(user: User, result: AssessmentResult) -> Tuple[bool, Optional[str]]:
        if user.is_staff or user.role == Role.ADMIN:
            return True, None

        if result.student_id != user.id:
            return False, "You are not authorized to view this result."

        if result.status != ResultStatus.FINALIZED:
            return False, "Result evaluation is still in progress."

        visibility = result.assessment.result_visibility
        now = timezone.now()

        if visibility == ResultVisibility.IMMEDIATE:
            return True, None
        elif visibility == ResultVisibility.AFTER_DEADLINE:
            if now < result.assessment.end_datetime:
                return False, f"Results will be available after the assessment deadline ({result.assessment.end_datetime.isoformat()})."
            return True, None
        elif visibility == ResultVisibility.MANUAL:
            if not result.is_released:
                return False, "Results have not been released by the administrator."
            return True, None

        return True, None


class ResultFinalizationService:
    """
    Authoritative Result Projection Service.
    Consumes the evaluation states from Phase 5 and Phase 6, computes decimal scores,
    and creates immutable AssessmentResult, QuestionResult, and HistoricalResultSummary records.
    """
    @classmethod
    @transaction.atomic
    def finalize_attempt(
        cls,
        attempt_id: str,
        actor: Optional[User] = None,
        request=None
    ) -> AssessmentResult:
        attempt = TestAttempt.objects.select_for_update().filter(id=attempt_id).select_related(
            'student', 'assessment', 'assessment_snapshot'
        ).first()

        if not attempt:
            raise NotFound("Test attempt not found.")

        # Ensure attempt is in terminal state
        if attempt.status not in [AttemptStatus.SUBMITTED, AttemptStatus.EXPIRED, AttemptStatus.CANCELLED]:
            raise DRFValidationError({"status": f"Cannot finalize attempt in {attempt.status} status."})

        # Fetch or initialize AssessmentResult
        result, created = AssessmentResult.objects.select_for_update().get_or_create(
            attempt=attempt,
            defaults={
                'student': attempt.student,
                'assessment': attempt.assessment,
                'assessment_snapshot': attempt.assessment_snapshot,
                'status': ResultStatus.PROCESSING
            }
        )

        if not created and result.status == ResultStatus.FINALIZED:
            # Idempotent return
            return result

        result.status = ResultStatus.PROCESSING
        result.save(update_fields=['status', 'updated_at'])

        snapshot = attempt.assessment_snapshot
        server_bundle = snapshot.server_evaluation_bundle or {}
        server_questions = server_bundle.get('questions_eval', server_bundle.get('questions', {}))
        snapshot_questions = {
            sq.snapshot_question_id: sq
            for sq in snapshot.snapshot_questions.all()
        }

        # Fetch candidate answers
        answers_map = {}
        for ans in attempt.answers.all():
            answers_map[ans.question_id] = ans
            if ans.snapshot_question:
                answers_map[ans.snapshot_question.snapshot_question_id] = ans
            answers_map[str(ans.snapshot_question_id)] = ans

        # Fetch authoritative coding submissions for this attempt
        coding_submissions_map = {}
        coding_subs = CodeSubmission.objects.filter(
            attempt=attempt,
            submission_type=SubmissionType.SUBMIT
        ).order_by('created_at')
        for cs in coding_subs:
            # Latest SUBMIT submission is authoritative per Phase 6 contract
            coding_submissions_map[str(cs.snapshot_question_id)] = cs
            if cs.snapshot_question:
                coding_submissions_map[cs.snapshot_question.snapshot_question_id] = cs

        total_earned = Decimal('0.00')
        total_possible = Decimal('0.00')
        correct_count = 0
        partially_correct_count = 0
        incorrect_count = 0
        skipped_count = 0
        answered_count = 0

        # Delete existing draft QuestionResults if retrying
        QuestionResult.objects.filter(assessment_result=result).delete()
        question_results_to_create = []

        for q_id, sq in snapshot_questions.items():
            eval_info = server_questions.get(q_id, {})
            max_pts = Decimal(str(sq.points)).quantize(Decimal('0.01'))
            total_possible += max_pts

            ans = answers_map.get(q_id) or answers_map.get(str(sq.id))
            earned_pts = Decimal('0.00')
            is_correct = False
            is_partially_correct = False
            is_skipped = True
            eval_details = {}

            q_type = sq.question_type

            if q_type in [QuestionType.MCQ, QuestionType.TRUE_FALSE]:
                if ans and ans.is_answered and ans.selected_options:
                    is_skipped = False
                    answered_count += 1
                    user_opts = ans.selected_options
                    correct_cfg = eval_info.get('correct_type_config', {})
                    correct_opts = correct_cfg.get('correct_options', [])
                    
                    # Normalize comparison
                    if set(map(str, user_opts)) == set(map(str, correct_opts)):
                        is_correct = True
                        earned_pts = max_pts
                        correct_count += 1
                    else:
                        incorrect_count += 1
                        if sq.negative_marking_enabled and sq.negative_points > 0:
                            earned_pts = -Decimal(str(sq.negative_points)).quantize(Decimal('0.01'))
                    eval_details = {
                        "user_selected": user_opts,
                        "is_correct": is_correct
                    }
                else:
                    skipped_count += 1

            elif q_type == QuestionType.MULTI_SELECT:
                if ans and ans.is_answered and ans.selected_options:
                    is_skipped = False
                    answered_count += 1
                    user_opts = set(map(str, ans.selected_options))
                    correct_opts = set(map(str, eval_info.get('correct_type_config', {}).get('correct_options', [])))

                    if user_opts == correct_opts:
                        is_correct = True
                        earned_pts = max_pts
                        correct_count += 1
                    else:
                        incorrect_count += 1
                        if sq.negative_marking_enabled and sq.negative_points > 0:
                            earned_pts = -Decimal(str(sq.negative_points)).quantize(Decimal('0.01'))
                    eval_details = {
                        "user_selected": list(user_opts),
                        "is_correct": is_correct
                    }
                else:
                    skipped_count += 1

            elif q_type == QuestionType.SHORT_ANSWER:
                if ans and ans.is_answered and ans.text_response:
                    is_skipped = False
                    answered_count += 1
                    user_text = ans.text_response.strip()
                    correct_cfg = eval_info.get('correct_type_config', {})
                    exact_matches = [m.strip() for m in correct_cfg.get('exact_matches', [])]
                    case_sensitive = correct_cfg.get('case_sensitive', False)

                    match_found = False
                    if case_sensitive:
                        match_found = user_text in exact_matches
                    else:
                        user_lower = user_text.lower()
                        match_found = any(user_lower == m.lower() for m in exact_matches)

                    if match_found:
                        is_correct = True
                        earned_pts = max_pts
                        correct_count += 1
                    else:
                        incorrect_count += 1
                        if sq.negative_marking_enabled and sq.negative_points > 0:
                            earned_pts = -Decimal(str(sq.negative_points)).quantize(Decimal('0.01'))
                    eval_details = {"user_text": user_text, "is_correct": is_correct}
                else:
                    skipped_count += 1

            elif q_type == QuestionType.CODING:
                # Consume Phase 6 authoritative CodeSubmission
                code_sub = coding_submissions_map.get(str(sq.id)) or coding_submissions_map.get(sq.snapshot_question_id)
                if code_sub and code_sub.status == 'COMPLETED':
                    is_skipped = False
                    answered_count += 1
                    earned_pts = Decimal(str(code_sub.score_awarded)).quantize(Decimal('0.01'))
                    
                    if earned_pts >= max_pts:
                        is_correct = True
                        correct_count += 1
                    elif earned_pts > Decimal('0.00'):
                        is_partially_correct = True
                        partially_correct_count += 1
                    else:
                        incorrect_count += 1

                    eval_details = {
                        "submission_id": str(code_sub.id),
                        "verdict": code_sub.verdict,
                        "passed_test_cases": code_sub.passed_test_cases,
                        "total_test_cases": code_sub.total_test_cases,
                        "execution_time_ms": code_sub.execution_time_ms,
                        "memory_used_kb": code_sub.memory_used_kb
                    }
                else:
                    skipped_count += 1

            elif q_type == QuestionType.SQL:
                # 1. Check for authoritative completed CodeSubmission
                sql_sub = coding_submissions_map.get(str(sq.id)) or coding_submissions_map.get(sq.snapshot_question_id)
                if sql_sub and sql_sub.status == 'COMPLETED':
                    is_skipped = False
                    answered_count += 1
                    earned_pts = Decimal(str(sql_sub.score_awarded)).quantize(Decimal('0.01'))
                    is_correct = (sql_sub.verdict == 'ACCEPTED')
                    if is_correct:
                        correct_count += 1
                    else:
                        incorrect_count += 1

                    eval_details = {
                        "submission_id": str(sql_sub.id),
                        "verdict": sql_sub.verdict,
                        "execution_time_ms": sql_sub.execution_time_ms,
                        "sql_query": sql_sub.source_code,
                        "is_correct": is_correct
                    }
                elif ans and ans.is_answered and ans.sql_response and ans.sql_response.strip():
                    is_skipped = False
                    answered_count += 1
                    candidate_sql = ans.sql_response.strip()

                    # Authoritative evaluation of student's autosaved SQL query
                    from apps.evaluator.sql_sandbox import SQLExecutionService
                    server_sql_eval = eval_info.get('server_sql_eval', {})
                    schema_setup = server_sql_eval.get('schema_setup_sql') or (sq.sql_config or {}).get('schema_setup_sql', '')
                    expected_def = server_sql_eval.get('expected_result_definition') or ''
                    time_limit_ms = server_sql_eval.get('time_limit_ms') or (sq.sql_config or {}).get('time_limit_ms', 3000)
                    ord_req = server_sql_eval.get('ordering_required')
                    if ord_req is None:
                        ord_req = (sq.sql_config or {}).get('ordering_required')

                    eval_res = SQLExecutionService.evaluate_query(
                        candidate_sql=candidate_sql,
                        schema_setup_sql=schema_setup,
                        expected_result_definition=expected_def,
                        time_limit_ms=time_limit_ms,
                        ordering_required=ord_req
                    )

                    is_correct = eval_res.get('is_correct', False)
                    if is_correct:
                        earned_pts = max_pts
                        correct_count += 1
                    else:
                        earned_pts = Decimal('0.00')
                        incorrect_count += 1

                    eval_details = {
                        "verdict": eval_res.get('verdict'),
                        "execution_time_ms": eval_res.get('execution_time_ms', 0),
                        "sql_query": candidate_sql,
                        "is_correct": is_correct,
                        "error_message": eval_res.get('error_message') if not is_correct else None
                    }
                else:
                    skipped_count += 1

            total_earned += earned_pts

            qr = QuestionResult(
                assessment_result=result,
                snapshot_question=sq,
                question_id=sq.snapshot_question_id,
                question_type=sq.question_type,
                earned_points=earned_pts,
                max_points=max_pts,
                is_correct=is_correct,
                is_partially_correct=is_partially_correct,
                is_skipped=is_skipped,
                evaluation_details=eval_details
            )
            question_results_to_create.append(qr)

        QuestionResult.objects.bulk_create(question_results_to_create)

        # Clamping score: Minimum 0.00
        final_earned = max(Decimal('0.00'), total_earned).quantize(Decimal('0.01'))
        percentage = Decimal('0.00')
        if total_possible > Decimal('0.00'):
            percentage = ((final_earned / total_possible) * Decimal('100.00')).quantize(Decimal('0.01'))

        # Determine pass/fail from frozen snapshot passing_percentage
        passing_pct = Decimal(str(snapshot.snapshot_data.get('passing_percentage', attempt.assessment.passing_percentage or 0.00)))
        is_passed = percentage >= passing_pct if passing_pct > 0 else True

        # Calculate time spent
        time_spent_seconds = 0
        if attempt.started_at and attempt.submitted_at:
            time_spent_seconds = int((attempt.submitted_at - attempt.started_at).total_seconds())

        result.total_score_earned = final_earned
        result.total_possible_score = total_possible.quantize(Decimal('0.01'))
        result.percentage = percentage
        result.is_passed = is_passed
        result.total_questions = len(snapshot_questions)
        result.answered_questions = answered_count
        result.correct_questions = correct_count
        result.partially_correct_questions = partially_correct_count
        result.incorrect_questions = incorrect_count
        result.skipped_questions = skipped_count
        result.time_spent_seconds = max(0, time_spent_seconds)
        result.finalized_at = timezone.now()
        result.status = ResultStatus.FINALIZED
        result.save()

        # Generate permanent HistoricalResultSummary in same transaction
        roll_no = getattr(attempt.student.student_profile, 'roll_number', '') if hasattr(attempt.student, 'student_profile') else ''
        euid = getattr(attempt.student.student_profile, 'euid', '') if hasattr(attempt.student, 'student_profile') else ''

        HistoricalResultSummary.objects.update_or_create(
            student=attempt.student,
            assessment_id=attempt.assessment_id,
            defaults={
                'student_euid': euid,
                'student_roll_number': roll_no,
                'assessment_snapshot_id': snapshot.id,
                'assessment_title_snapshot': snapshot.snapshot_data.get('title', attempt.assessment.title),
                'total_score_earned': final_earned,
                'total_possible_score': total_possible,
                'percentage': percentage,
                'is_passed': is_passed,
                'completion_status': attempt.status,
                'started_at': attempt.started_at or timezone.now(),
                'completed_at': attempt.submitted_at or timezone.now(),
                'details_purged': False
            }
        )

        AuditService.log(
            action="ASSESSMENT_RESULT_FINALIZED",
            actor=actor or attempt.student,
            target_type="AssessmentResult",
            target_id=str(result.id),
            metadata={
                "attempt_id": str(attempt.id),
                "assessment_id": str(attempt.assessment_id),
                "total_score_earned": str(final_earned),
                "percentage": str(percentage),
                "is_passed": is_passed
            },
            request=request
        )

        # Phase 9: Bind RetentionRecord 1:1 with finalized attempt
        try:
            from apps.retention.services import RetentionPolicyEngine
            RetentionPolicyEngine.create_retention_record_for_finalized_attempt(attempt)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"Could not bind RetentionRecord on finalization: {e}")

        # Notify attempt channel if Channels is active
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            channel_layer = get_channel_layer()
            if channel_layer:
                async_to_sync(channel_layer.group_send)(
                    f"attempt_{attempt.id}",
                    {
                        "type": "assessment_result_finalized",
                        "result_id": str(result.id),
                        "status": "FINALIZED"
                    }
                )
        except Exception:
            pass

        return result


class AnalyticsService:
    """
    Computes statistical and performance analytics across assessments, questions,
    students, topics, and proctoring risks.
    """
    @staticmethod
    def get_assessment_analytics(assessment_id: str) -> Dict[str, Any]:
        assessment = Assessment.objects.filter(id=assessment_id).first()
        if not assessment:
            raise NotFound("Assessment not found.")

        results = AssessmentResult.objects.filter(
            assessment=assessment,
            status=ResultStatus.FINALIZED
        ).select_related('attempt', 'student')

        total_assigned = assessment.assignments.count()
        total_started = assessment.attempts.count()
        total_completed = results.count()
        total_expired = assessment.attempts.filter(status=AttemptStatus.EXPIRED).count()

        if total_completed == 0:
            return {
                "assessment_id": str(assessment.id),
                "assessment_title": assessment.title,
                "cohort_metrics": {
                    "total_assigned": total_assigned,
                    "total_started": total_started,
                    "total_completed": 0,
                    "total_expired": total_expired,
                    "completion_rate_percentage": 0.0,
                    "pass_rate_percentage": 0.0
                },
                "score_statistics": {
                    "mean_score": "0.00",
                    "median_score": "0.00",
                    "highest_score": "0.00",
                    "lowest_score": "0.00",
                    "standard_deviation": "0.00",
                    "quartiles": {"q1": "0.00", "q2": "0.00", "q3": "0.00"}
                },
                "score_distribution": [],
                "proctoring_risk_correlation": {"is_available": False, "reason": "No finalized attempts"}
            }

        scores = sorted([float(r.total_score_earned) for r in results])
        n = len(scores)
        mean_score = sum(scores) / n
        highest_score = max(scores)
        lowest_score = min(scores)

        # Median & Quartiles
        def _get_percentile(data, p):
            k = (len(data) - 1) * p
            f = math.floor(k)
            c = math.ceil(k)
            if f == c:
                return data[int(k)]
            d0 = data[int(f)] * (c - k)
            d1 = data[int(c)] * (k - f)
            return d0 + d1

        median_score = _get_percentile(scores, 0.50)
        q1 = _get_percentile(scores, 0.25)
        q3 = _get_percentile(scores, 0.75)

        # Standard Deviation
        variance = sum((x - mean_score) ** 2 for x in scores) / n
        std_dev = math.sqrt(variance)

        # Histogram Distribution (10-point buckets)
        buckets = [0] * 10
        for s in scores:
            idx = min(9, int(s // 10))
            buckets[idx] += 1

        distribution = [
            {"bucket": f"{i*10}-{(i+1)*10}", "count": buckets[i]}
            for i in range(10)
        ]

        # Pass rate
        passed_count = sum(1 for r in results if r.is_passed)
        pass_rate = round((passed_count / n) * 100.0, 2)
        completion_rate = round((total_completed / max(1, total_started)) * 100.0, 2)

        # Proctoring risk correlation with N >= 10 privacy threshold
        proctoring_data = {"is_available": False}
        if n >= 10:
            risk_dist = {"NORMAL": {"count": 0, "total_score": 0.0},
                         "LOW": {"count": 0, "total_score": 0.0},
                         "MEDIUM": {"count": 0, "total_score": 0.0},
                         "HIGH": {"count": 0, "total_score": 0.0},
                         "CRITICAL": {"count": 0, "total_score": 0.0}}
            
            proct_sessions = ProctoringSession.objects.filter(
                attempt__assessment=assessment
            ).select_related('attempt')
            
            sess_map = {ps.attempt_id: ps for ps in proct_sessions}

            for r in results:
                ps = sess_map.get(r.attempt_id)
                band = ps.risk_band if ps else RiskBand.NORMAL
                if band in risk_dist:
                    risk_dist[band]["count"] += 1
                    risk_dist[band]["total_score"] += float(r.total_score_earned)

            dist_out = {}
            for k, v in risk_dist.items():
                avg = round(v["total_score"] / v["count"], 2) if v["count"] > 0 else 0.0
                dist_out[k] = {"count": v["count"], "average_score": str(avg)}

            proctoring_data = {
                "is_available": True,
                "min_cohort_threshold_met": True,
                "distribution": dist_out
            }
        else:
            proctoring_data = {
                "is_available": False,
                "reason": "Cohort size below privacy threshold (N < 10)"
            }

        return {
            "assessment_id": str(assessment.id),
            "assessment_title": assessment.title,
            "cohort_metrics": {
                "total_assigned": total_assigned,
                "total_started": total_started,
                "total_completed": total_completed,
                "total_expired": total_expired,
                "completion_rate_percentage": completion_rate,
                "pass_rate_percentage": pass_rate
            },
            "score_statistics": {
                "mean_score": str(round(mean_score, 2)),
                "median_score": str(round(median_score, 2)),
                "highest_score": str(round(highest_score, 2)),
                "lowest_score": str(round(lowest_score, 2)),
                "standard_deviation": str(round(std_dev, 2)),
                "quartiles": {
                    "q1": str(round(q1, 2)),
                    "q2": str(round(median_score, 2)),
                    "q3": str(round(q3, 2))
                }
            },
            "score_distribution": distribution,
            "proctoring_risk_correlation": proctoring_data
        }

    @staticmethod
    def get_question_analytics(assessment_id: str) -> List[Dict[str, Any]]:
        assessment = Assessment.objects.filter(id=assessment_id).first()
        if not assessment:
            raise NotFound("Assessment not found.")

        q_results = QuestionResult.objects.filter(
            assessment_result__assessment=assessment,
            assessment_result__status=ResultStatus.FINALIZED
        ).select_related('snapshot_question', 'assessment_result')

        # Group by snapshot_question
        from collections import defaultdict
        grouped = defaultdict(list)
        for qr in q_results:
            grouped[qr.snapshot_question].append(qr)

        out = []
        for sq, results in grouped.items():
            total = len(results)
            if total == 0:
                continue

            correct = sum(1 for r in results if r.is_correct)
            partial = sum(1 for r in results if r.is_partially_correct)
            incorrect = sum(1 for r in results if not r.is_correct and not r.is_partially_correct and not r.is_skipped)
            skipped = sum(1 for r in results if r.is_skipped)
            
            difficulty_p = round(correct / total, 2)
            avg_score = round(sum(float(r.earned_points) for r in results) / total, 2)
            avg_time = round(sum(r.time_spent_seconds for r in results) / total, 1)

            # Discrimination Index (Upper 27% vs Lower 27%) if N >= 10
            discrimination_d = None
            if total >= 10:
                sorted_results = sorted(results, key=lambda x: float(x.assessment_result.total_score_earned))
                k = int(math.ceil(total * 0.27))
                lower_grp = sorted_results[:k]
                upper_grp = sorted_results[-k:]
                p_upper = sum(1 for r in upper_grp if r.is_correct) / k
                p_lower = sum(1 for r in lower_grp if r.is_correct) / k
                discrimination_d = round(p_upper - p_lower, 2)

            out.append({
                "snapshot_question_id": sq.snapshot_question_id,
                "order": sq.order,
                "question_type": sq.question_type,
                "title": sq.title,
                "difficulty_index_p": difficulty_p,
                "discrimination_index_d": discrimination_d,
                "average_score": str(avg_score),
                "max_points": str(sq.points),
                "average_time_spent_seconds": avg_time,
                "breakdown": {
                    "total_responses": total,
                    "correct": correct,
                    "partially_correct": partial,
                    "incorrect": incorrect,
                    "skipped": skipped
                }
            })

        out.sort(key=lambda x: x['order'])
        return out

    @staticmethod
    def get_student_topic_analytics(student: User) -> List[Dict[str, Any]]:
        q_results = QuestionResult.objects.filter(
            assessment_result__student=student,
            assessment_result__status=ResultStatus.FINALIZED
        ).select_related('snapshot_question')

        tag_aggregates = {}
        for qr in q_results:
            tags = qr.snapshot_question.tags or []
            for tag in tags:
                if tag not in tag_aggregates:
                    tag_aggregates[tag] = {
                        "tag_name": tag,
                        "questions_attempted": 0,
                        "earned_points": Decimal('0.00'),
                        "max_points": Decimal('0.00'),
                    }
                tag_aggregates[tag]["questions_attempted"] += 1
                tag_aggregates[tag]["earned_points"] += qr.earned_points
                tag_aggregates[tag]["max_points"] += qr.max_points

        out = []
        for tag, data in tag_aggregates.items():
            acc = 0.0
            if data["max_points"] > Decimal('0.00'):
                acc = float((data["earned_points"] / data["max_points"]) * Decimal('100.00'))
            out.append({
                "tag_name": tag,
                "questions_attempted": data["questions_attempted"],
                "accuracy_percentage": round(acc, 2),
                "earned_points": str(data["earned_points"]),
                "max_points": str(data["max_points"])
            })

        out.sort(key=lambda x: x['tag_name'])
        return out


class ReportService:
    """
    Asynchronous Report Generation & Controlled Export Engine.
    Renders PDF (ReportLab), XLSX (openpyxl), and CSV (pandas) exports with
    Formula Injection sanitization and SHA-256 integrity digests.
    """
    REPORTS_DIR = os.path.join(settings.BASE_DIR, 'media', 'reports')

    @classmethod
    def create_report_job(
        cls,
        user: User,
        report_type: str,
        format: str,
        assessment_id: Optional[str] = None,
        student_id: Optional[str] = None
    ) -> ReportJob:
        assessment = None
        if assessment_id:
            assessment = Assessment.objects.filter(id=assessment_id).first()
            if not assessment:
                raise NotFound("Assessment not found.")

        target_student = None
        if student_id:
            target_student = User.objects.filter(id=student_id).first()

        # Authorization checks
        if user.role == Role.STUDENT:
            if report_type != ReportType.STUDENT_SCORECARD:
                raise PermissionDenied("Students can only generate personal scorecards.")
            target_student = user

        job = ReportJob.objects.create(
            requested_by=user,
            report_type=report_type,
            format=format,
            status=ReportStatus.PENDING,
            assessment=assessment,
            student=target_student,
            expires_at=timezone.now() + timedelta(days=7)
        )

        from .tasks import generate_report_job_task
        generate_report_job_task.delay(str(job.id))
        return job

    @classmethod
    def generate_report(cls, job_id: str) -> ReportJob:
        job = ReportJob.objects.filter(id=job_id).select_related('assessment', 'student', 'requested_by').first()
        if not job:
            raise NotFound("Report job not found.")

        job.status = ReportStatus.PROCESSING
        job.save(update_fields=['status', 'updated_at'])

        os.makedirs(cls.REPORTS_DIR, exist_ok=True)
        ext = job.format.lower()
        file_name = f"{job.id}.{ext}"
        file_path = os.path.join(cls.REPORTS_DIR, file_name)

        try:
            if job.format == ReportFormat.CSV:
                cls._generate_csv(job, file_path)
            elif job.format == ReportFormat.XLSX:
                cls._generate_xlsx(job, file_path)
            elif job.format == ReportFormat.PDF:
                cls._generate_pdf(job, file_path)

            file_size = os.path.getsize(file_path)
            with open(file_path, 'rb') as f:
                sha256 = hashlib.sha256(f.read()).hexdigest()

            job.file_path = file_path
            job.file_size_bytes = file_size
            job.sha256_hash = sha256
            job.status = ReportStatus.COMPLETED
            job.completed_at = timezone.now()
            job.save()

            AuditService.log(
                action="REPORT_GENERATED",
                actor=job.requested_by,
                target_type="ReportJob",
                target_id=str(job.id),
                metadata={
                    "report_type": job.report_type,
                    "format": job.format,
                    "file_size": file_size,
                    "sha256": sha256
                }
            )
            return job

        except Exception as e:
            job.status = ReportStatus.FAILED
            job.error_message = str(e)[:500]
            job.save(update_fields=['status', 'error_message', 'updated_at'])
            raise

    @classmethod
    def _generate_csv(cls, job: ReportJob, file_path: str):
        """Controlled CSV Export Schema (No raw table dumps)"""
        results = AssessmentResult.objects.filter(
            assessment=job.assessment,
            status=ResultStatus.FINALIZED
        ).select_related('student', 'student__student_profile', 'attempt')

        rows = []
        for r in results:
            prof = getattr(r.student, 'student_profile', None)
            euid = getattr(prof, 'euid', '') if prof else ''
            roll = getattr(prof, 'roll_number', '') if prof else ''
            name = r.student.email

            rows.append({
                "euid": _sanitize_formula_injection(euid),
                "roll_number": _sanitize_formula_injection(roll),
                "student_name": _sanitize_formula_injection(name),
                "student_email": _sanitize_formula_injection(r.student.email),
                "assessment_id": str(job.assessment.id),
                "assessment_title": _sanitize_formula_injection(job.assessment.title),
                "total_score_earned": str(r.total_score_earned),
                "total_possible_score": str(r.total_possible_score),
                "percentage": str(r.percentage),
                "is_passed": "PASS" if r.is_passed else "FAIL",
                "started_at": r.attempt.started_at.isoformat() if r.attempt.started_at else "",
                "submitted_at": r.attempt.submitted_at.isoformat() if r.attempt.submitted_at else "",
                "time_spent_seconds": r.time_spent_seconds,
            })

        df = pd.DataFrame(rows)
        df.to_csv(file_path, index=False)

    @classmethod
    def _generate_xlsx(cls, job: ReportJob, file_path: str):
        """Excel Gradebook with Deterministic Precomputed Cell Values"""
        wb = openpyxl.Workbook()
        ws_roster = wb.active
        ws_roster.title = "Candidate Roster"

        headers = ["EUID", "Roll Number", "Name", "Email", "Score Earned", "Max Points", "Percentage", "Verdict", "Time Spent (s)", "Submitted At"]
        ws_roster.append([_sanitize_formula_injection(h) for h in headers])

        # Style header
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws_roster.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        results = AssessmentResult.objects.filter(
            assessment=job.assessment,
            status=ResultStatus.FINALIZED
        ).select_related('student', 'student__student_profile', 'attempt')

        for r in results:
            prof = getattr(r.student, 'student_profile', None)
            euid = getattr(prof, 'euid', '') if prof else ''
            roll = getattr(prof, 'roll_number', '') if prof else ''
            name = r.student.email

            row = [
                _sanitize_formula_injection(euid),
                _sanitize_formula_injection(roll),
                _sanitize_formula_injection(name),
                _sanitize_formula_injection(r.student.email),
                float(r.total_score_earned),
                float(r.total_possible_score),
                float(r.percentage),
                "PASS" if r.is_passed else "FAIL",
                r.time_spent_seconds,
                r.attempt.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if r.attempt.submitted_at else ""
            ]
            ws_roster.append(row)

        # Sheet 2: Question Item Analysis
        ws_questions = wb.create_sheet(title="Question Analysis")
        q_headers = ["Order", "Title", "Type", "Max Points", "Difficulty (P)", "Avg Score"]
        ws_questions.append([_sanitize_formula_injection(h) for h in q_headers])
        for col in range(1, len(q_headers) + 1):
            cell = ws_questions.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        q_analytics = AnalyticsService.get_question_analytics(str(job.assessment.id))
        for qa in q_analytics:
            ws_questions.append([
                qa['order'],
                _sanitize_formula_injection(qa['title']),
                qa['question_type'],
                float(qa['max_points']),
                qa['difficulty_index_p'],
                float(qa['average_score'])
            ])

        wb.save(file_path)

    @classmethod
    def _generate_pdf(cls, job: ReportJob, file_path: str):
        """ReportLab PDF Candidate Scorecard / Executive Summary"""
        doc = SimpleDocTemplate(file_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=12
        )
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=18
        )

        if job.report_type == ReportType.STUDENT_SCORECARD:
            res = AssessmentResult.objects.filter(
                assessment=job.assessment,
                student=job.student,
                status=ResultStatus.FINALIZED
            ).select_related('student', 'student__student_profile', 'assessment').first()

            if not res:
                raise NotFound("Student assessment result not found for PDF export.")

            prof = getattr(res.student, 'student_profile', None)
            euid = getattr(prof, 'euid', 'N/A') if prof else 'N/A'
            roll = getattr(prof, 'roll_number', 'N/A') if prof else 'N/A'

            story.append(Paragraph("CODEGUARD Assessment Scorecard", title_style))
            story.append(Paragraph(f"Assessment: {res.assessment.title}", subtitle_style))
            story.append(Spacer(1, 10))

            meta_data = [
                ["Candidate Email:", res.student.email, "EUID:", euid],
                ["Roll Number:", roll, "Email:", res.student.email],
                ["Total Score:", f"{res.total_score_earned} / {res.total_possible_score}", "Percentage:", f"{res.percentage}%"],
                ["Verdict:", "PASSED" if res.is_passed else "FAILED", "Time Spent:", f"{res.time_spent_seconds} seconds"]
            ]
            t_meta = Table(meta_data, colWidths=[110, 160, 90, 160])
            t_meta.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#0F172A')),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ]))
            story.append(t_meta)
            story.append(Spacer(1, 20))

            story.append(Paragraph("Question Scoring Breakdown", styles['Heading2']))
            q_rows = [["#", "Type", "Points Earned", "Max Points", "Status"]]
            for qr in res.question_results.all().order_by('snapshot_question__order'):
                status_str = "Correct" if qr.is_correct else ("Partial" if qr.is_partially_correct else ("Skipped" if qr.is_skipped else "Incorrect"))
                q_rows.append([
                    str(qr.snapshot_question.order),
                    qr.question_type,
                    str(qr.earned_points),
                    str(qr.max_points),
                    status_str
                ])

            t_q = Table(q_rows, colWidths=[40, 120, 120, 120, 120])
            t_q.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
            ]))
            story.append(t_q)

        else:
            # Assessment Summary
            analytics = AnalyticsService.get_assessment_analytics(str(job.assessment.id))
            story.append(Paragraph("CODEGUARD Assessment Executive Summary", title_style))
            story.append(Paragraph(f"Assessment: {job.assessment.title}", subtitle_style))
            story.append(Spacer(1, 10))

            cohort = analytics['cohort_metrics']
            stats = analytics['score_statistics']
            summary_data = [
                ["Total Assigned:", str(cohort['total_assigned']), "Completed:", str(cohort['total_completed'])],
                ["Pass Rate:", f"{cohort['pass_rate_percentage']}%", "Completion Rate:", f"{cohort['completion_rate_percentage']}%"],
                ["Mean Score:", stats['mean_score'], "Median Score:", stats['median_score']],
                ["Highest Score:", stats['highest_score'], "Lowest Score:", stats['lowest_score']],
                ["Std Deviation:", stats['standard_deviation'], "IQR (Q1 - Q3):", f"{stats['quartiles']['q1']} - {stats['quartiles']['q3']}"]
            ]
            t_sum = Table(summary_data, colWidths=[120, 140, 120, 140])
            t_sum.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6)
            ]))
            story.append(t_sum)

        doc.build(story)


class RetentionService:
    """
    Guarantees the Retention / Finalization Synchronization Invariant:
    A terminal attempt MUST complete result finalization and generate its
    HistoricalResultSummary before detailed source data is eligible for purging.
    """
    @classmethod
    def is_eligible_for_purge(cls, attempt_id: str) -> bool:
        attempt = TestAttempt.objects.filter(id=attempt_id).first()
        if not attempt:
            return False

        # 1. Attempt must be in terminal state
        if attempt.status not in [AttemptStatus.SUBMITTED, AttemptStatus.EXPIRED, AttemptStatus.CANCELLED]:
            return False

        # 2. Result must exist and be FINALIZED
        result = AssessmentResult.objects.filter(attempt=attempt).first()
        if not result or result.status != ResultStatus.FINALIZED:
            return False

        # 3. Permanent HistoricalResultSummary must exist
        has_summary = HistoricalResultSummary.objects.filter(
            student=attempt.student,
            assessment_id=attempt.assessment_id
        ).exists()
        if not has_summary:
            return False

        # 4. Check if active in-flight report jobs exist for this assessment
        active_reports = ReportJob.objects.filter(
            assessment=attempt.assessment,
            status=ReportStatus.PROCESSING
        ).exists()
        if active_reports:
            return False

        return True

    @classmethod
    def purge_detailed_attempt_data(cls, attempt_id: str, force: bool = False) -> bool:
        """
        Scrubs detailed attempt answers and raw telemetry while preserving
        AssessmentResult header and permanent HistoricalResultSummary.
        """
        if not force and not cls.is_eligible_for_purge(attempt_id):
            logger.warning(f"Attempt {attempt_id} is NOT eligible for detailed data purge.")
            return False

        with transaction.atomic():
            attempt = TestAttempt.objects.select_for_update().filter(id=attempt_id).first()
            if not attempt:
                return False

            # Delete detailed attempt answers (Phase 5)
            attempt.answers.all().delete()

            # Mark HistoricalResultSummary as details_purged
            HistoricalResultSummary.objects.filter(
                student=attempt.student,
                assessment_id=attempt.assessment_id
            ).update(details_purged=True)

            logger.info(f"Successfully purged detailed telemetry for attempt {attempt_id}")
            return True

