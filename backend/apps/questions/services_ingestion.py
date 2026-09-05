import io
import os
import re
import csv
import uuid
import json
import logging
import subprocess
from typing import Dict, Any, List, Tuple, Optional
import shutil
from pathlib import Path

import pandas as pd
from PIL import Image, ImageOps, ImageEnhance
from django.conf import settings
from django.core.exceptions import ValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

from apps.accounts.models import User
from .models import (
    Question,
    QuestionVersion,
    QuestionType,
    Difficulty,
    VersionStatus,
    CodingLanguage,
    Tag,
)
from .services import QuestionService

logger = logging.getLogger(__name__)

TEMP_IMAGE_DIR = getattr(settings, 'MEDIA_ROOT', Path('/tmp')) / 'temp_question_images'
OCR_HELPER_PATH = Path(__file__).resolve().parent / 'bin' / 'ocr_helper'


class SpreadsheetQuestionImporter:
    """
    Service responsible for validating, parsing, and previewing Excel (.xlsx)
    and CSV (.csv) spreadsheets containing question authoring definitions,
    and creating draft questions following Admin confirmation.
    """

    SUPPORTED_EXTENSIONS = ('.csv', '.xlsx', '.xls')

    COLUMN_MAPPING = {
        'question_title': 'title',
        'title': 'title',
        'question_type': 'question_type',
        'type': 'question_type',
        'difficulty': 'difficulty',
        'total_points': 'points',
        'points': 'points',
        'tags': 'tags',
        'problem_statement': 'problem_statement',
        'description': 'problem_statement',
        'student_instructions': 'instructions',
        'instructions': 'instructions',
        'option_a': 'option_a',
        'option_b': 'option_b',
        'option_c': 'option_c',
        'option_d': 'option_d',
        'correct_option': 'correct_option',
        'correct_answer': 'correct_option',
        'language': 'languages',
        'languages': 'languages',
        'allowed_languages': 'languages',
        'starter_code': 'starter_code',
        'constraints': 'constraints',
        'sample_input': 'sample_input',
        'sample_output': 'sample_output',
        'sample_points': 'sample_points',
        'dialect': 'allowed_dialect',
        'schema_setup_sql': 'schema_setup_sql',
    }

    @classmethod
    def generate_template_csv(cls) -> bytes:
        """
        Generates official CSV template containing sample rows for MCQ and Coding questions.
        """
        headers = [
            'question_title',
            'question_type',
            'difficulty',
            'total_points',
            'tags',
            'problem_statement',
            'student_instructions',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'correct_option',
            'allowed_languages',
            'starter_code',
            'constraints',
            'sample_input',
            'sample_output',
            'sample_points'
        ]

        rows = [
            [
                'Time Complexity of Hash Map Lookup',
                'MCQ',
                'EASY',
                '10',
                'Data Structures, Hash Table',
                'What is the average time complexity of looking up a key in a well-distributed hash table?',
                'Select the single correct Big-O complexity.',
                'O(1)',
                'O(n)',
                'O(log n)',
                'O(n^2)',
                'A',
                '',
                '',
                '',
                '',
                '',
                ''
            ],
            [
                'Reverse an Array In-Place',
                'CODING',
                'EASY',
                '20',
                'Arrays, Algorithms, Two Pointers',
                'Given an array of integers, reverse the elements in-place without allocating extra array space.',
                'Return the array or modify it in place.',
                '',
                '',
                '',
                '',
                '',
                'PYTHON,CPP,JAVA',
                'def reverse_array(arr):\n    # Write your solution here\n    pass',
                '1 <= len(arr) <= 10^5\n-10^9 <= arr[i] <= 10^9',
                '1 2 3 4 5',
                '5 4 3 2 1',
                '10'
            ],
            [
                'Identify Valid HTTP Status Codes',
                'MULTI_SELECT',
                'MEDIUM',
                '15',
                'Networking, Web',
                'Which of the following are valid 4xx Client Error HTTP status codes?',
                'Select all that apply.',
                '400 Bad Request',
                '404 Not Found',
                '500 Internal Server Error',
                '403 Forbidden',
                'A,B,D',
                '',
                '',
                '',
                '',
                '',
                ''
            ]
        ]

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(rows)
        return buffer.getvalue().encode('utf-8')

    @classmethod
    def generate_template_xlsx(cls) -> bytes:
        """
        Generates official Excel (.xlsx) template with formatted columns and examples.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CODEGUARD Questions Template"

        headers = [
            'question_title',
            'question_type',
            'difficulty',
            'total_points',
            'tags',
            'problem_statement',
            'student_instructions',
            'option_a',
            'option_b',
            'option_c',
            'option_d',
            'correct_option',
            'allowed_languages',
            'starter_code',
            'constraints',
            'sample_input',
            'sample_output',
            'sample_points'
        ]

        ws.append(headers)

        sample_rows = [
            [
                'Time Complexity of Hash Map Lookup',
                'MCQ',
                'EASY',
                10,
                'Data Structures, Hash Table',
                'What is the average time complexity of looking up a key in a well-distributed hash table?',
                'Select the single correct Big-O complexity.',
                'O(1)',
                'O(n)',
                'O(log n)',
                'O(n^2)',
                'A',
                '',
                '',
                '',
                '',
                '',
                ''
            ],
            [
                'Reverse an Array In-Place',
                'CODING',
                'EASY',
                20,
                'Arrays, Algorithms, Two Pointers',
                'Given an array of integers, reverse the elements in-place without allocating extra array space.',
                'Return the array or modify it in place.',
                '',
                '',
                '',
                '',
                '',
                'PYTHON,CPP,JAVA',
                'def reverse_array(arr):\n    # Write your solution here\n    pass',
                '1 <= len(arr) <= 10^5\n-10^9 <= arr[i] <= 10^9',
                '1 2 3 4 5',
                '5 4 3 2 1',
                10
            ]
        ]

        for r in sample_rows:
            ws.append(r)

        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(18, len(headers[col_idx - 1]) + 4)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @classmethod
    def parse_and_validate_spreadsheet(cls, file_obj, filename: str) -> Dict[str, Any]:
        """
        Parses spreadsheet into row-level items, executes validation invariants,
        detects potential duplicates against existing question titles, and produces a structured preview.
        """
        ext = Path(filename).suffix.lower()
        if ext not in cls.SUPPORTED_EXTENSIONS:
            raise DRFValidationError(f"Unsupported spreadsheet format '{ext}'. Supported formats: .csv, .xlsx, .xls")

        try:
            if ext == '.csv':
                df = pd.read_csv(file_obj, keep_default_na=False)
            else:
                df = pd.read_excel(file_obj, keep_default_na=False)
        except Exception as e:
            logger.warning(f"Spreadsheet parse failure: {e}")
            raise DRFValidationError("Unable to read spreadsheet. Ensure the file is not corrupted and contains valid table data.")

        if df.empty:
            raise DRFValidationError("The uploaded spreadsheet contains no rows.")

        # Normalize column names
        normalized_cols = {}
        for col in df.columns:
            cleaned = str(col).strip().lower().replace(' ', '_').replace('-', '_')
            target_key = cls.COLUMN_MAPPING.get(cleaned, cleaned)
            normalized_cols[col] = target_key
        df.rename(columns=normalized_cols, inplace=True)

        parsed_rows = []
        valid_count = 0
        error_count = 0
        duplicate_count = 0

        # Pre-fetch existing active question titles for fast duplicate detection
        existing_titles = set(QuestionVersion.objects.values_list('title', flat=True))
        normalized_existing_titles = {t.strip().lower(): t for t in existing_titles if t}

        valid_types = {t[0] for t in QuestionType.choices}
        valid_diffs = {d[0] for d in Difficulty.choices}
        valid_languages = {l[0] for l in CodingLanguage.choices}

        for idx, row in df.iterrows():
            row_num = idx + 2  # 1-indexed (header is 1, data starts row 2)
            row_errors: List[str] = []
            is_duplicate = False

            title = str(row.get('title', '')).strip()
            if not title:
                row_errors.append("Missing required field: question_title")

            raw_type = str(row.get('question_type', '')).strip().upper()
            if not raw_type:
                row_errors.append("Missing required field: question_type")
            elif raw_type not in valid_types:
                row_errors.append(f"Invalid question_type '{raw_type}'. Supported: {', '.join(sorted(valid_types))}")

            raw_diff = str(row.get('difficulty', 'MEDIUM')).strip().upper() or 'MEDIUM'
            if raw_diff not in valid_diffs:
                row_errors.append(f"Invalid difficulty '{raw_diff}'. Supported: EASY, MEDIUM, HARD")

            raw_points = row.get('points', 10)
            try:
                points = int(raw_points) if raw_points != '' else 10
                if points < 1:
                    row_errors.append("Points must be an integer greater than or equal to 1")
            except (ValueError, TypeError):
                row_errors.append(f"Invalid total_points '{raw_points}'. Must be an integer")
                points = 10

            description = str(row.get('problem_statement', '')).strip()
            if not description:
                row_errors.append("Missing required field: problem_statement / description")

            instructions = str(row.get('instructions', '')).strip()

            raw_tags = str(row.get('tags', '')).strip()
            tags = [t.strip() for t in raw_tags.split(',') if t.strip()] if raw_tags else []

            # Question type specific validation
            type_config: Dict[str, Any] = {'_source': 'SPREADSHEET_IMPORT'}
            coding_config: Optional[Dict[str, Any]] = None
            sql_config: Optional[Dict[str, Any]] = None
            test_cases: List[Dict[str, Any]] = []

            if raw_type == QuestionType.MCQ:
                opt_a = str(row.get('option_a', '')).strip()
                opt_b = str(row.get('option_b', '')).strip()
                opt_c = str(row.get('option_c', '')).strip()
                opt_d = str(row.get('option_d', '')).strip()
                correct_opt = str(row.get('correct_option', '')).strip().upper()

                options = []
                for key, val in [('A', opt_a), ('B', opt_b), ('C', opt_c), ('D', opt_d)]:
                    if val:
                        options.append({'id': key, 'text': val})

                if len(options) < 2:
                    row_errors.append("MCQ questions require at least 2 options (option_a and option_b)")
                if not correct_opt:
                    row_errors.append("MCQ questions require a correct_option (e.g. A, B, C, or D)")
                elif correct_opt not in [o['id'] for o in options]:
                    row_errors.append(f"correct_option '{correct_opt}' does not match any provided options")

                type_config['options'] = options
                type_config['correct_option'] = correct_opt

            elif raw_type == QuestionType.MULTI_SELECT:
                opt_a = str(row.get('option_a', '')).strip()
                opt_b = str(row.get('option_b', '')).strip()
                opt_c = str(row.get('option_c', '')).strip()
                opt_d = str(row.get('option_d', '')).strip()
                correct_opts = [c.strip().upper() for c in str(row.get('correct_option', '')).split(',') if c.strip()]

                options = []
                for key, val in [('A', opt_a), ('B', opt_b), ('C', opt_c), ('D', opt_d)]:
                    if val:
                        options.append({'id': key, 'text': val})

                if len(options) < 2:
                    row_errors.append("Multi-select questions require at least 2 options")
                if not correct_opts:
                    row_errors.append("Multi-select questions require at least one correct_option")
                else:
                    avail_ids = {o['id'] for o in options}
                    for co in correct_opts:
                        if co not in avail_ids:
                            row_errors.append(f"correct_option '{co}' is not among provided options")

                type_config['options'] = options
                type_config['correct_options'] = correct_opts

            elif raw_type == QuestionType.TRUE_FALSE:
                ans = str(row.get('correct_option', '')).strip().upper()
                if ans not in ('TRUE', 'FALSE', 'T', 'F'):
                    row_errors.append("TRUE_FALSE questions require correct_option to be TRUE or FALSE")
                type_config['correct_answer'] = True if ans in ('TRUE', 'T') else False

            elif raw_type == QuestionType.SHORT_ANSWER:
                ans = str(row.get('correct_option', '')).strip()
                if not ans:
                    row_errors.append("SHORT_ANSWER questions require a reference correct_option / answer")
                type_config['correct_answer'] = ans

            elif raw_type == QuestionType.CODING:
                raw_langs = str(row.get('languages', '')).strip().upper()
                if raw_langs:
                    langs = [l.strip() for l in raw_langs.split(',') if l.strip()]
                    invalid_langs = [l for l in langs if l not in valid_languages]
                    if invalid_langs:
                        row_errors.append(f"Invalid coding language(s): {', '.join(invalid_langs)}. Supported: PYTHON, CPP, JAVA")
                else:
                    langs = [CodingLanguage.PYTHON, CodingLanguage.CPP, CodingLanguage.JAVA]

                starter_code = str(row.get('starter_code', ''))
                constraints = str(row.get('constraints', '')).strip()

                coding_config = {
                    'problem_statement': description,
                    'input_description': '',
                    'output_description': '',
                    'constraints': constraints,
                    'allowed_languages': langs,
                    'starter_code': starter_code,
                    'time_limit_ms': 2000,
                    'memory_limit_mb': 256,
                }

                sample_in = str(row.get('sample_input', '')).strip()
                sample_out = str(row.get('sample_output', '')).strip()
                if sample_in or sample_out:
                    sample_pts = int(row.get('sample_points', 5) or 5)
                    test_cases.append({
                        'input_data': sample_in,
                        'expected_output': sample_out,
                        'points': sample_pts,
                        'is_hidden': False,
                        'execution_order': 1
                    })

            elif raw_type == QuestionType.SQL:
                dialect = str(row.get('allowed_dialect', 'MYSQL')).strip().upper() or 'MYSQL'
                schema_sql = str(row.get('schema_setup_sql', '')).strip()
                sql_config = {
                    'problem_statement': description,
                    'schema_setup_sql': schema_sql,
                    'expected_result_definition': '',
                    'allowed_dialect': dialect,
                    'time_limit_ms': 3000,
                }

            # Duplicate detection check
            if title and title.lower() in normalized_existing_titles:
                is_duplicate = True
                duplicate_count += 1

            if row_errors:
                status = "ERROR"
                error_count += 1
            elif is_duplicate:
                status = "DUPLICATE_WARNING"
                valid_count += 1
            else:
                status = "VALID"
                valid_count += 1

            parsed_rows.append({
                'row_number': row_num,
                'status': status,
                'is_duplicate': is_duplicate,
                'duplicate_of': normalized_existing_titles.get(title.lower(), "") if is_duplicate else None,
                'errors': row_errors,
                'data': {
                    'title': title,
                    'question_type': raw_type,
                    'difficulty': raw_diff,
                    'points': points,
                    'tags': tags,
                    'description': description,
                    'instructions': instructions,
                    'type_config': type_config,
                    'coding_config': coding_config,
                    'sql_config': sql_config,
                    'test_cases': test_cases,
                }
            })

        return {
            'total_rows': len(parsed_rows),
            'valid_count': valid_count,
            'error_count': error_count,
            'duplicate_count': duplicate_count,
            'rows': parsed_rows
        }

    @classmethod
    def commit_imported_rows(
        cls,
        rows_data: List[Dict[str, Any]],
        actor: User,
        request=None
    ) -> Dict[str, Any]:
        """
        Commits valid spreadsheet rows to Question and initial QuestionVersion entities.
        Each question is created with status DRAFT. Never publishes automatically.
        """
        created_questions = []
        skipped_rows = []

        for row_entry in rows_data:
            data = row_entry.get('data') or row_entry
            title = data.get('title', '').strip()
            q_type = data.get('question_type', '').strip().upper()
            diff = data.get('difficulty', 'MEDIUM').strip().upper()
            points = int(data.get('points', 10))
            desc = data.get('description', '').strip()
            instructions = data.get('instructions', '').strip()
            tags = data.get('tags', [])
            type_config = data.get('type_config', {})
            coding_config = data.get('coding_config')
            sql_config = data.get('sql_config')
            test_cases = data.get('test_cases', [])

            if not title or not q_type or not desc:
                skipped_rows.append({'title': title, 'reason': 'Missing required fields'})
                continue

            try:
                question, version = QuestionService.create_question(
                    question_type=q_type,
                    title=title,
                    description=desc,
                    instructions=instructions,
                    points=points,
                    difficulty=diff,
                    tags=tags,
                    type_config=type_config,
                    coding_config_data=coding_config,
                    test_cases_data=test_cases if test_cases else None,
                    sql_config_data=sql_config,
                    actor=actor,
                    request=request
                )
                created_questions.append({
                    'question_id': str(question.id),
                    'version_id': str(version.id),
                    'title': version.title,
                    'question_type': version.question_type,
                    'difficulty': version.difficulty,
                    'points': version.points,
                    'status': version.status,  # Will always be DRAFT
                })
            except Exception as e:
                logger.error(f"Failed to create draft question '{title}': {e}")
                skipped_rows.append({'title': title, 'reason': str(e)})

        return {
            'total_submitted': len(rows_data),
            'created_count': len(created_questions),
            'skipped_count': len(skipped_rows),
            'created_questions': created_questions,
            'skipped_rows': skipped_rows
        }


class BaseOCREngine:
    """Base interface for optical character recognition engines."""
    def extract_text(self, image_path: Path) -> List[str]:
        raise NotImplementedError


class AppleVisionOCREngine(BaseOCREngine):
    """Native macOS Vision framework engine using compiled Swift helper."""
    def extract_text(self, image_path: Path) -> List[str]:
        if OCR_HELPER_PATH.exists() and os.access(OCR_HELPER_PATH, os.X_OK):
            try:
                res = subprocess.run(
                    [str(OCR_HELPER_PATH), str(image_path)],
                    capture_output=True,
                    text=True,
                    timeout=20
                )
                if res.returncode == 0 and res.stdout.strip():
                    lines = json.loads(res.stdout)
                    if isinstance(lines, list):
                        return [str(l).strip() for l in lines if str(l).strip()]
            except Exception as e:
                logger.warning(f"Native Apple Vision OCR invocation failed: {e}")
        return []


class TesseractOCREngine(BaseOCREngine):
    """Linux/Docker compatible Tesseract OCR engine."""
    def extract_text(self, image_path: Path) -> List[str]:
        tesseract_bin = shutil.which('tesseract')
        if tesseract_bin:
            try:
                res = subprocess.run(
                    [tesseract_bin, str(image_path), 'stdout', '--oem', '1'],
                    capture_output=True,
                    text=True,
                    timeout=20
                )
                if res.returncode == 0 and res.stdout.strip():
                    return [line.strip() for line in res.stdout.splitlines() if line.strip()]
            except Exception as e:
                logger.warning(f"Tesseract OCR engine failed: {e}")
        return []


class ImageQuestionExtractor:
    """
    Service responsible for secure image upload validation, executing native
    or containerized OCR extraction, parsing layout structures, and returning
    an editable extraction draft for mandatory Admin review.
    """

    ALLOWED_MIME_TYPES = {'image/png', 'image/jpeg', 'image/webp'}
    ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp'}
    MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024  # 10MB
    MAX_IMAGE_PIXELS = 25000000  # Decompression bomb safety (25MP)

    @classmethod
    def get_ocr_engine(cls) -> Tuple[BaseOCREngine, str]:
        """Selects the best available OCR engine for the host platform."""
        if OCR_HELPER_PATH.exists() and os.access(OCR_HELPER_PATH, os.X_OK):
            return AppleVisionOCREngine(), "Apple Vision Native"
        if shutil.which('tesseract'):
            return TesseractOCREngine(), "Tesseract OCR"
        return AppleVisionOCREngine(), "None (Fallback)"

    @classmethod
    def validate_and_store_image(cls, file_obj, original_filename: str) -> Tuple[str, Path]:
        """
        Validates image file size, extension, MIME type, and image header safety.
        Protects against decompression bombs and malformed image payloads.
        Stores temporarily in media cache with UUID filename.
        """
        ext = Path(original_filename).suffix.lower()
        if ext not in cls.ALLOWED_EXTENSIONS:
            raise DRFValidationError(f"Unsupported image extension '{ext}'. Supported: .png, .jpg, .jpeg, .webp")

        size = file_obj.size if hasattr(file_obj, 'size') else len(file_obj.read())
        if hasattr(file_obj, 'seek'):
            file_obj.seek(0)

        if size > cls.MAX_FILE_SIZE_BYTES:
            raise DRFValidationError("Uploaded image exceeds the maximum allowed limit of 10 MB.")

        # Pillow decompression bomb protection & integrity check
        Image.MAX_IMAGE_PIXELS = cls.MAX_IMAGE_PIXELS
        try:
            with Image.open(file_obj) as img:
                if img.format not in ('PNG', 'JPEG', 'WEBP'):
                    raise DRFValidationError(f"Invalid image format '{img.format}'. Only PNG, JPEG, and WEBP are supported.")
                w, h = img.size
                if w * h > cls.MAX_IMAGE_PIXELS:
                    raise DRFValidationError("Image pixel dimensions exceed safe decompression limits.")
                img.verify()

            if hasattr(file_obj, 'seek'):
                file_obj.seek(0)
        except DRFValidationError:
            raise
        except Exception as e:
            logger.warning(f"Invalid image content: {e}")
            raise DRFValidationError("The uploaded file is corrupt or not a readable image.")

        TEMP_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
        unique_name = f"{uuid.uuid4().hex}{ext}"
        destination_path = TEMP_IMAGE_DIR / unique_name

        with open(destination_path, 'wb+') as dest:
            for chunk in file_obj.chunks() if hasattr(file_obj, 'chunks') else [file_obj.read()]:
                dest.write(chunk)

        return unique_name, destination_path

    @classmethod
    def preprocess_image(cls, original_path: Path) -> Path:
        """
        Preprocesses image with Pillow to maximize OCR readability:
        - Auto-orientates based on EXIF
        - Normalizes RGB channels
        - Moderately enhances contrast for sharp glyph boundaries
        """
        prep_path = original_path.parent / f"prep_{original_path.name}.png"
        try:
            with Image.open(original_path) as img:
                img = ImageOps.exif_transpose(img) or img
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                enhancer = ImageEnhance.Contrast(img)
                enhanced = enhancer.enhance(1.3)
                enhanced.save(prep_path, format='PNG')
                return prep_path
        except Exception as e:
            logger.warning(f"Pillow preprocessing failed, using original: {e}")
            return original_path

    @classmethod
    def extract_text_from_image(cls, image_path: Path) -> List[str]:
        """
        Preprocesses the image then invokes the platform-appropriate OCR engine.
        Ensures temporary preprocessed images are cleaned up.
        """
        prep_path = cls.preprocess_image(image_path)
        engine, engine_name = cls.get_ocr_engine()
        logger.info(f"Extracting OCR text using engine: {engine_name}")

        try:
            lines = engine.extract_text(prep_path)
        finally:
            if prep_path != image_path and prep_path.exists():
                try:
                    prep_path.unlink()
                except Exception:
                    pass

        return lines

    @classmethod
    def map_extracted_lines_to_question_structure(
        cls,
        lines: List[str],
        image_id: str
    ) -> Dict[str, Any]:
        """
        Maps recognized text lines into standard CODEGUARD Question configuration schema.
        Detects question type, title, problem statement, constraints, input/output formats,
        examples, test cases, and starter code.
        """
        raw_text = "\n".join(lines).strip()

        detected_type = QuestionType.CODING
        title = "Extracted Question"
        difficulty = Difficulty.MEDIUM
        points = 10
        tags = ["Extracted"]
        instructions = ""
        type_config: Dict[str, Any] = {'_source': 'IMAGE_EXTRACTION'}
        coding_config: Optional[Dict[str, Any]] = None
        sql_config: Optional[Dict[str, Any]] = None
        examples: List[Dict[str, Any]] = []
        test_cases: List[Dict[str, Any]] = []
        warnings: List[str] = [
            "Extracted content is created in DRAFT status. Human review and verification are required before publishing."
        ]

        if not lines:
            return {
                'image_id': image_id,
                'status': 'DRAFT',
                'question_type': QuestionType.CODING,
                'title': 'New Extracted Question',
                'difficulty': 'MEDIUM',
                'points': 10,
                'description': 'Please enter the question details from your uploaded image.',
                'instructions': '',
                'tags': ['Extracted'],
                'type_config': type_config,
                'coding_config': {
                    'problem_statement': '',
                    'allowed_languages': ['PYTHON', 'CPP', 'JAVA'],
                    'starter_code': '',
                    'constraints': '',
                    'input_format': '',
                    'output_format': '',
                    'time_limit_ms': 2000,
                    'memory_limit_mb': 256
                },
                'examples': [],
                'test_cases': [],
                'confidence_score': 0,
                'confidence_level': 'LOW',
                'confidence_notice': 'No clear text could be automatically extracted from the image. Please use the original image on the left panel to complete the fields manually.',
                'review_warnings': warnings
            }

        # 1. Title heuristic: First meaningful non-trivial line
        for line in lines:
            cleaned = line.strip()
            if cleaned and len(cleaned) > 3 and not cleaned.lower().startswith(('example', 'given', 'problem', 'question', 'note', 'input', 'output')):
                title = cleaned[:100]
                break

        # 2. Type Classification
        mcq_signals = [re.match(r'^[A-Da-d][\.\)]\s+', l.strip()) for l in lines]
        coding_signals = any(
            kw in raw_text for kw in ['def ', 'class ', 'public static', '#include', 'import ', 'return ', 'void ', 'int ']
        ) or any(
            kw in raw_text.lower() for kw in ['time complexity', 'space complexity', 'constraints', 'input format', 'sample input']
        )
        sql_signals = any(kw in raw_text.upper() for kw in ['SELECT ', 'FROM ', 'WHERE ', 'CREATE TABLE', 'INSERT INTO'])

        if sql_signals and not coding_signals:
            detected_type = QuestionType.SQL
        elif any(mcq_signals):
            detected_type = QuestionType.MCQ
        else:
            detected_type = QuestionType.CODING

        # 3. Structure Parsing
        if detected_type == QuestionType.MCQ:
            options = []
            option_pattern = re.compile(r'^([A-Da-d])[\.\)]\s+(.*)$')
            problem_lines = []

            for line in lines:
                m = option_pattern.match(line.strip())
                if m:
                    opt_id = m.group(1).upper()
                    opt_text = m.group(2).strip()
                    options.append({'id': opt_id, 'text': opt_text})
                else:
                    problem_lines.append(line)

            description = "\n".join(problem_lines).strip()
            type_config['options'] = options
            type_config['correct_option'] = options[0]['id'] if options else 'A'

        elif detected_type == QuestionType.CODING:
            problem_parts = []
            constraint_parts = []
            input_format_parts = []
            output_format_parts = []
            code_parts = []

            current_section = 'problem'
            ex_counter = 1
            cur_example: Optional[Dict[str, str]] = None

            for line in lines:
                lower = line.lower().strip()

                # Detect section switches
                if lower.startswith(('constraint', 'constraints:')):
                    current_section = 'constraints'
                    continue
                elif lower.startswith(('input format', 'input:')):
                    if 'example' in lower or current_section == 'example':
                        pass  # handled in example block
                    else:
                        current_section = 'input_format'
                        continue
                elif lower.startswith(('output format', 'output:')):
                    if 'example' in lower or current_section == 'example':
                        pass
                    else:
                        current_section = 'output_format'
                        continue
                elif lower.startswith(('starter code', 'code:')):
                    current_section = 'code'
                    continue
                elif 'example' in lower and ('1' in lower or '2' in lower or 'example:' in lower):
                    current_section = 'example'
                    if cur_example:
                        examples.append(cur_example)
                    cur_example = {'name': f'Example {ex_counter}', 'input': '', 'output': '', 'explanation': ''}
                    ex_counter += 1
                    continue

                if current_section == 'example' and cur_example:
                    if lower.startswith('input:'):
                        cur_example['input'] = line.split(':', 1)[1].strip()
                    elif lower.startswith('output:'):
                        cur_example['output'] = line.split(':', 1)[1].strip()
                    elif lower.startswith('explanation:'):
                        cur_example['explanation'] = line.split(':', 1)[1].strip()
                    elif not cur_example['output']:
                        if cur_example['input']:
                            cur_example['input'] += f"\n{line}"
                    else:
                        cur_example['explanation'] += f" {line}"
                elif current_section == 'constraints':
                    constraint_parts.append(line)
                elif current_section == 'input_format':
                    input_format_parts.append(line)
                elif current_section == 'output_format':
                    output_format_parts.append(line)
                elif current_section == 'code' or any(line.strip().startswith(kw) for kw in ['def ', 'class ', '#include', 'int ']):
                    code_parts.append(line)
                else:
                    problem_parts.append(line)

            if cur_example:
                examples.append(cur_example)

            description = "\n".join(problem_parts).strip() or raw_text
            constraints = "\n".join(constraint_parts).strip()
            starter_code = "\n".join(code_parts).strip()
            input_format = "\n".join(input_format_parts).strip()
            output_format = "\n".join(output_format_parts).strip()

            # Map parsed examples to sample test cases
            for idx, ex in enumerate(examples, start=1):
                if ex.get('input') or ex.get('output'):
                    test_cases.append({
                        'name': f"Sample Case {idx}",
                        'input': ex.get('input', ''),
                        'expected_output': ex.get('output', ''),
                        'is_sample': True,
                        'is_hidden': False,
                        'points': 5,
                        'explanation': ex.get('explanation', '')
                    })

            coding_config = {
                'problem_statement': description,
                'allowed_languages': ['PYTHON', 'CPP', 'JAVA'],
                'starter_code': starter_code,
                'constraints': constraints,
                'input_format': input_format,
                'output_format': output_format,
                'time_limit_ms': 2000,
                'memory_limit_mb': 256
            }

        elif detected_type == QuestionType.SQL:
            sql_config = {
                'problem_statement': raw_text,
                'schema_setup_sql': '-- Extracted schema setup\n',
                'expected_result_definition': '',
                'allowed_dialect': 'MYSQL',
                'time_limit_ms': 3000
            }
            description = raw_text

        # Confidence Score Calculation
        score = 20
        if len(lines) >= 5:
            score += 30
        if detected_type == QuestionType.CODING and (examples or coding_config.get('constraints')):
            score += 30
        if coding_config and coding_config.get('starter_code'):
            score += 15
        if title and title != "Extracted Question":
            score += 5
        score = min(100, score)

        level = 'HIGH' if score >= 80 else ('MEDIUM' if score >= 40 else 'LOW')

        return {
            'image_id': image_id,
            'status': 'DRAFT',
            'question_type': detected_type,
            'title': title,
            'difficulty': difficulty,
            'points': points,
            'description': description,
            'instructions': instructions,
            'tags': tags,
            'type_config': type_config,
            'coding_config': coding_config,
            'sql_config': sql_config,
            'examples': examples,
            'test_cases': test_cases,
            'confidence_score': score,
            'confidence_level': level,
            'confidence_notice': f"OCR completed with {level} confidence ({score}%). Review all extracted fields before saving.",
            'review_warnings': warnings
        }
