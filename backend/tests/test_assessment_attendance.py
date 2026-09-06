import pytest
import io
import openpyxl
from datetime import timedelta
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from apps.accounts.models import User, Role, Section, StudentProfile
from apps.assessments.models import (
    Assessment,
    AssessmentStatus,
    AssessmentAssignment,
    AssignmentStatus,
    TestAttempt,
    AttemptStatus,
    AssessmentSnapshot,
)
from apps.assessments.services import AssessmentAttendanceService, AssessmentService


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def admin_user(db):
    return User.objects.create_user(
        email='admin_attend@test.com',
        password='Password123!',
        role=Role.ADMIN,
        is_active=True
    )


@pytest.fixture
def student_user_1(db):
    user = User.objects.create_user(
        email='student1_attend@test.com',
        password='Password123!',
        display_name='Alice Smith',
        role=Role.STUDENT,
        is_active=True
    )
    section = Section.objects.create(code='SEC-A', name='Section Alpha', is_active=True)
    StudentProfile.objects.create(
        user=user,
        roll_number='ROLL-001',
        euid='EUID-001',
        section=section
    )
    return user


@pytest.fixture
def student_user_2(db):
    user = User.objects.create_user(
        email='student2_attend@test.com',
        password='Password123!',
        display_name='Bob Jones',
        role=Role.STUDENT,
        is_active=True
    )
    section = Section.objects.create(code='SEC-B', name='Section Beta', is_active=True)
    StudentProfile.objects.create(
        user=user,
        roll_number='ROLL-002',
        euid='EUID-002',
        section=section
    )
    return user


@pytest.fixture
def student_user_3(db):
    user = User.objects.create_user(
        email='student3_attend@test.com',
        password='Password123!',
        display_name='Charlie Brown',
        role=Role.STUDENT,
        is_active=True
    )
    # Direct / no section
    StudentProfile.objects.create(
        user=user,
        roll_number='ROLL-003',
        euid='EUID-003',
        section=None
    )
    return user


@pytest.fixture
def published_assessment(db, admin_user):
    now = timezone.now()
    return Assessment.objects.create(
        title='Data Structures Midterm',
        description='Comprehensive DS Exam',
        start_datetime=now - timedelta(hours=2),
        end_datetime=now + timedelta(hours=2),
        duration_minutes=60,
        total_points=100,
        created_by=admin_user,
        status=AssessmentStatus.PUBLISHED
    )


@pytest.fixture
def snapshot(db, published_assessment):
    return AssessmentSnapshot.objects.create(
        assessment=published_assessment,
        version_number=1,
        snapshot_data={},
        server_evaluation_bundle={}
    )


@pytest.mark.django_db
class TestAssessmentAttendanceService:

    def test_zero_assigned_attendance(self, published_assessment):
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_assigned'] == 0
        assert summary['total_attended'] == 0
        assert summary['total_not_attended'] == 0
        assert summary['attendance_percentage'] == 0.0
        assert len(data['results']) == 0

    def test_assigned_with_no_attempts_is_not_attended(
        self, published_assessment, student_user_1, admin_user
    ):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_assigned'] == 1
        assert summary['total_attended'] == 0
        assert summary['total_not_attended'] == 1
        assert summary['attendance_percentage'] == 0.0

        row = data['results'][0]
        assert row['attendance_status'] == 'NOT_ATTENDED'
        assert row['attempt_status'] == AttemptStatus.NOT_STARTED

    def test_started_attempt_is_attended(
        self, published_assessment, snapshot, student_user_1, admin_user
    ):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        now = timezone.now()
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.IN_PROGRESS,
            started_at=now - timedelta(minutes=15)
        )
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_assigned'] == 1
        assert summary['total_attended'] == 1
        assert summary['total_not_attended'] == 0
        assert summary['total_in_progress'] == 1
        assert summary['attendance_percentage'] == 100.0

        row = data['results'][0]
        assert row['attendance_status'] == 'ATTENDED'
        assert row['attempt_status'] == AttemptStatus.IN_PROGRESS

    def test_submitted_attempt_counts_in_summary(
        self, published_assessment, snapshot, student_user_1, admin_user
    ):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        now = timezone.now()
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.SUBMITTED,
            started_at=now - timedelta(minutes=45),
            submitted_at=now
        )
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_attended'] == 1
        assert summary['total_submitted'] == 1
        assert summary['total_in_progress'] == 0

        row = data['results'][0]
        assert row['attendance_status'] == 'ATTENDED'
        assert row['attempt_status'] == AttemptStatus.SUBMITTED
        assert row['duration_seconds'] == 45 * 60

    def test_cancelled_attempt_after_start_is_attended(
        self, published_assessment, snapshot, student_user_1, admin_user
    ):
        """
        Invariant 7.2: If started_at is NOT NULL and status is CANCELLED:
        Attendance = ATTENDED, Attempt Status = CANCELLED.
        """
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        now = timezone.now()
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.CANCELLED,
            started_at=now - timedelta(minutes=20)
        )
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_attended'] == 1
        row = data['results'][0]
        assert row['attendance_status'] == 'ATTENDED'
        assert row['attempt_status'] == AttemptStatus.CANCELLED

    def test_cancelled_attempt_before_start_is_not_attended(
        self, published_assessment, snapshot, student_user_1, admin_user
    ):
        """
        Invariant 7.2: If started_at is NULL and status is CANCELLED:
        Attendance = NOT_ATTENDED.
        """
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.CANCELLED,
            started_at=None
        )
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_attended'] == 0
        assert summary['total_not_attended'] == 1
        row = data['results'][0]
        assert row['attendance_status'] == 'NOT_ATTENDED'
        assert row['attempt_status'] == AttemptStatus.CANCELLED

    def test_multiple_attempts_count_student_once(
        self, published_assessment, snapshot, student_user_1, admin_user
    ):
        """
        Invariant 7.3: Attendance MUST count students, not attempts.
        Attempt 1: SUBMITTED
        Attempt 2: IN_PROGRESS
        Student counted once as ATTENDED, latest attempt status = IN_PROGRESS.
        """
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        now = timezone.now()
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.SUBMITTED,
            started_at=now - timedelta(hours=1),
            submitted_at=now - timedelta(minutes=30)
        )
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=2,
            status=AttemptStatus.IN_PROGRESS,
            started_at=now - timedelta(minutes=10)
        )
        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_assigned'] == 1
        assert summary['total_attended'] == 1
        assert summary['total_in_progress'] == 1
        assert len(data['results']) == 1
        row = data['results'][0]
        assert row['attendance_status'] == 'ATTENDED'
        assert row['attempt_status'] == AttemptStatus.IN_PROGRESS
        assert row['attempts_count'] == 2

    def test_section_breakdown_and_direct_student(
        self, published_assessment, snapshot, student_user_1, student_user_2, student_user_3, admin_user
    ):
        """
        Section reporting uses StudentProfile.section.
        student 1 in SEC-A, student 2 in SEC-B, student 3 has no section.
        """
        for s in [student_user_1, student_user_2, student_user_3]:
            AssessmentAssignment.objects.create(
                assessment=published_assessment,
                student=s,
                assigned_by=admin_user,
                status=AssignmentStatus.ASSIGNED
            )

        now = timezone.now()
        # Student 1 attended
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.SUBMITTED,
            started_at=now - timedelta(minutes=30),
            submitted_at=now
        )

        data = AssessmentAttendanceService.get_attendance_data(published_assessment)
        summary = data['summary']
        assert summary['total_assigned'] == 3
        assert summary['total_attended'] == 1
        assert summary['total_not_attended'] == 2
        assert summary['attendance_percentage'] == 33.3

        sections = {sec['section_name']: sec for sec in data['sections']}
        assert 'Section Alpha' in sections
        assert sections['Section Alpha']['assigned'] == 1
        assert sections['Section Alpha']['attended'] == 1
        assert sections['Section Alpha']['attendance_percentage'] == 100.0

        assert 'Section Beta' in sections
        assert sections['Section Beta']['assigned'] == 1
        assert sections['Section Beta']['attended'] == 0
        assert sections['Section Beta']['attendance_percentage'] == 0.0

        assert 'Unassigned' in sections
        assert sections['Unassigned']['assigned'] == 1
        assert sections['Unassigned']['attended'] == 0

    def test_pre_exam_detection(self, admin_user):
        """
        Invariant 7.5: Assessments starting in the future have is_pre_exam = True.
        """
        future = timezone.now() + timedelta(days=2)
        future_assessment = Assessment.objects.create(
            title='Future Exam',
            start_datetime=future,
            end_datetime=future + timedelta(hours=3),
            duration_minutes=90,
            total_points=50,
            created_by=admin_user,
            status=AssessmentStatus.PUBLISHED
        )
        data = AssessmentAttendanceService.get_attendance_data(future_assessment)
        assert data['summary']['is_pre_exam'] is True
        assert data['assessment']['is_pre_exam'] is True

    def test_pagination_does_not_corrupt_summary(
        self, published_assessment, snapshot, student_user_1, student_user_2, student_user_3, admin_user
    ):
        """
        Invariant 9.1: Summary statistics MUST NOT depend on pagination page_size.
        """
        for s in [student_user_1, student_user_2, student_user_3]:
            AssessmentAssignment.objects.create(
                assessment=published_assessment,
                student=s,
                assigned_by=admin_user,
                status=AssignmentStatus.ASSIGNED
            )
        now = timezone.now()
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.SUBMITTED,
            started_at=now - timedelta(minutes=20),
            submitted_at=now
        )

        # page_size=1
        data_p1 = AssessmentAttendanceService.get_attendance_data(
            published_assessment, page=1, page_size=1
        )
        assert data_p1['summary']['total_assigned'] == 3
        assert data_p1['summary']['total_attended'] == 1
        assert data_p1['total_count'] == 3
        assert len(data_p1['results']) == 1

        # page 2
        data_p2 = AssessmentAttendanceService.get_attendance_data(
            published_assessment, page=2, page_size=1
        )
        assert data_p2['summary']['total_assigned'] == 3
        assert data_p2['summary']['total_attended'] == 1
        assert len(data_p2['results']) == 1

    def test_search_and_filters(
        self, published_assessment, snapshot, student_user_1, student_user_2, admin_user
    ):
        for s in [student_user_1, student_user_2]:
            AssessmentAssignment.objects.create(
                assessment=published_assessment,
                student=s,
                assigned_by=admin_user,
                status=AssignmentStatus.ASSIGNED
            )
        now = timezone.now()
        TestAttempt.objects.create(
            assessment=published_assessment,
            assessment_snapshot=snapshot,
            student=student_user_1,
            attempt_number=1,
            status=AttemptStatus.SUBMITTED,
            started_at=now - timedelta(minutes=20),
            submitted_at=now
        )

        # Search for Alice
        data_alice = AssessmentAttendanceService.get_attendance_data(
            published_assessment, filters={'search': 'alice'}
        )
        assert data_alice['summary']['total_assigned'] == 1
        assert data_alice['results'][0]['student_name'] == 'Alice Smith'

        # Filter by attendance_status = ATTENDED
        data_att = AssessmentAttendanceService.get_attendance_data(
            published_assessment, filters={'attendance_status': 'ATTENDED'}
        )
        assert data_att['summary']['total_assigned'] == 1
        assert data_att['results'][0]['student_id'] == str(student_user_1.id)

        # Filter by attendance_status = NOT_ATTENDED
        data_not = AssessmentAttendanceService.get_attendance_data(
            published_assessment, filters={'attendance_status': 'NOT_ATTENDED'}
        )
        assert data_not['summary']['total_assigned'] == 1
        assert data_not['results'][0]['student_id'] == str(student_user_2.id)

    def test_xlsx_export(self, published_assessment, snapshot, student_user_1, admin_user):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        buf = AssessmentAttendanceService.export_attendance_xlsx(published_assessment)
        assert isinstance(buf, io.BytesIO)
        wb = openpyxl.load_workbook(buf)
        assert "Attendance Roster" in wb.sheetnames
        assert "Section Summary" in wb.sheetnames
        ws = wb["Attendance Roster"]
        assert ws.cell(row=2, column=1).value == "Alice Smith"

    def test_pdf_export(self, published_assessment, snapshot, student_user_1, admin_user):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        buf = AssessmentAttendanceService.export_attendance_pdf(published_assessment)
        assert isinstance(buf, io.BytesIO)
        content = buf.getvalue()
        assert content.startswith(b'%PDF')


@pytest.mark.django_db
class TestAttendanceAPIAndSecurity:

    def test_admin_can_access_attendance_api(
        self, api_client, admin_user, published_assessment, student_user_1
    ):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        api_client.force_authenticate(user=admin_user)
        res = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/')
        assert res.status_code == status.HTTP_200_OK
        assert res.data['status'] == 'success'
        assert res.data['data']['summary']['total_assigned'] == 1

    def test_student_cannot_access_attendance_api(
        self, api_client, student_user_1, published_assessment
    ):
        api_client.force_authenticate(user=student_user_1)
        res = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/')
        assert res.status_code == status.HTTP_403_FORBIDDEN

    def test_student_cannot_export_attendance(
        self, api_client, student_user_1, published_assessment
    ):
        api_client.force_authenticate(user=student_user_1)
        res = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/export/')
        assert res.status_code == status.HTTP_403_FORBIDDEN

    def test_unauthenticated_cannot_access_attendance(
        self, api_client, published_assessment
    ):
        res = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/')
        assert res.status_code == status.HTTP_401_UNAUTHORIZED

    def test_admin_can_export_xlsx(
        self, api_client, admin_user, published_assessment, student_user_1
    ):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        api_client.force_authenticate(user=admin_user)
        res = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/export/?format=xlsx')
        assert res.status_code == status.HTTP_200_OK
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in res['Content-Type']
        assert f'attendance_{published_assessment.id}.xlsx' in res['Content-Disposition']

    def test_admin_can_export_pdf(
        self, api_client, admin_user, published_assessment, student_user_1
    ):
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        api_client.force_authenticate(user=admin_user)
        res = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/export/?format=pdf')
        assert res.status_code == status.HTTP_200_OK
        assert 'application/pdf' in res['Content-Type']
        assert f'attendance_{published_assessment.id}.pdf' in res['Content-Disposition']

    def test_cross_assessment_data_isolation(
        self, api_client, admin_user, published_assessment, snapshot, student_user_1, student_user_2
    ):
        """
        Data from Assessment B must not leak into Assessment A attendance.
        """
        now = timezone.now()
        other_assessment = Assessment.objects.create(
            title='Other Assessment',
            start_datetime=now - timedelta(hours=1),
            end_datetime=now + timedelta(hours=1),
            duration_minutes=30,
            total_points=50,
            created_by=admin_user,
            status=AssessmentStatus.PUBLISHED
        )
        other_snapshot = AssessmentSnapshot.objects.create(
            assessment=other_assessment,
            version_number=1,
            snapshot_data={},
            server_evaluation_bundle={}
        )
        # Assign student 1 to assessment A, student 2 to assessment B
        AssessmentAssignment.objects.create(
            assessment=published_assessment,
            student=student_user_1,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        AssessmentAssignment.objects.create(
            assessment=other_assessment,
            student=student_user_2,
            assigned_by=admin_user,
            status=AssignmentStatus.ASSIGNED
        )
        TestAttempt.objects.create(
            assessment=other_assessment,
            assessment_snapshot=other_snapshot,
            student=student_user_2,
            attempt_number=1,
            status=AttemptStatus.SUBMITTED,
            started_at=now - timedelta(minutes=20),
            submitted_at=now
        )

        api_client.force_authenticate(user=admin_user)
        res_a = api_client.get(f'/api/v1/admin/assessments/{published_assessment.id}/attendance/')
        assert res_a.status_code == status.HTTP_200_OK
        assert res_a.data['data']['summary']['total_assigned'] == 1
        assert res_a.data['data']['results'][0]['student_id'] == str(student_user_1.id)
        assert res_a.data['data']['summary']['total_attended'] == 0


@pytest.mark.django_db
class TestSaveDraftAndPublishHardenings:

    def test_save_draft_with_audience_persists_atomically(
        self, api_client, admin_user, student_user_1, student_user_2
    ):
        """
        Save draft with target_section_ids and target_student_ids persists audience.
        """
        section = student_user_1.student_profile.section
        api_client.force_authenticate(user=admin_user)

        now = timezone.now()
        draft = Assessment.objects.create(
            title='Draft Exam',
            start_datetime=now + timedelta(days=1),
            end_datetime=now + timedelta(days=2),
            duration_minutes=60,
            total_points=50,
            created_by=admin_user,
            status=AssessmentStatus.DRAFT
        )

        payload = {
            "title": "Updated Draft Exam",
            "target_section_ids": [str(section.id)],
            "target_student_ids": [str(student_user_2.id)],
        }
        res = api_client.patch(f'/api/v1/admin/assessments/{draft.id}/', payload, format='json')
        assert res.status_code == status.HTTP_200_OK

        draft.refresh_from_db()
        assert draft.title == "Updated Draft Exam"
        assert draft.target_sections.filter(id=section.id).exists()
        assert draft.target_students.filter(id=student_user_2.id).exists()

    def test_save_draft_with_invalid_student_id_rolls_back(
        self, api_client, admin_user
    ):
        api_client.force_authenticate(user=admin_user)
        now = timezone.now()
        draft = Assessment.objects.create(
            title='Original Title',
            start_datetime=now + timedelta(days=1),
            end_datetime=now + timedelta(days=2),
            duration_minutes=60,
            total_points=50,
            created_by=admin_user,
            status=AssessmentStatus.DRAFT
        )

        fake_id = '00000000-0000-0000-0000-000000000099'
        payload = {
            "title": "Should Not Persist",
            "target_student_ids": [fake_id]
        }
        res = api_client.patch(f'/api/v1/admin/assessments/{draft.id}/', payload, format='json')
        assert res.status_code == status.HTTP_400_BAD_REQUEST

        draft.refresh_from_db()
        assert draft.title == "Original Title"

    def test_delete_draft_endpoint(
        self, api_client, admin_user, published_assessment
    ):
        api_client.force_authenticate(user=admin_user)

        # Published assessment cannot be deleted
        res = api_client.delete(f'/api/v1/admin/assessments/{published_assessment.id}/')
        assert res.status_code == status.HTTP_400_BAD_REQUEST
        assert Assessment.objects.filter(id=published_assessment.id).exists()

        # Draft assessment can be deleted
        now = timezone.now()
        draft = Assessment.objects.create(
            title='Deletable Draft',
            start_datetime=now + timedelta(days=1),
            end_datetime=now + timedelta(days=2),
            duration_minutes=60,
            total_points=50,
            created_by=admin_user,
            status=AssessmentStatus.DRAFT
        )
        res_del = api_client.delete(f'/api/v1/admin/assessments/{draft.id}/')
        assert res_del.status_code == status.HTTP_200_OK
        assert not Assessment.objects.filter(id=draft.id).exists()

    def test_double_publish_rejected(
        self, api_client, admin_user, published_assessment
    ):
        api_client.force_authenticate(user=admin_user)
        res = api_client.post(f'/api/v1/admin/assessments/{published_assessment.id}/publish/')
        assert res.status_code == status.HTTP_400_BAD_REQUEST
